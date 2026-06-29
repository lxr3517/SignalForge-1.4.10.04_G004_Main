from __future__ import annotations
from pathlib import Path
import shutil
import json
import pandas as pd
import numpy as np
import math
import re
import html
from fastapi import APIRouter, Depends, Form, Request, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.config import TEMPLATES_DIR, UPLOADS_DIR, PROJECTS_DIR, APP_NAME, APP_VERSION
from app.db import get_db
from app.schemas import ProjectCreate
from app.services.project_service import create_project, update_project, delete_project, list_projects, get_project, create_run, update_run_status, list_runs, get_run
from app.services.file_service import save_upload, load_dataframe, get_sheet_names
from app.services.quality_service import run_quality_checks
from app.services.auto_fix_engine import run_auto_fix_engine, load_auto_fix_audit
from app.services.guidance_service import build_quality_guidance, build_forecast_guidance
from app.services.forecast_service import ForecastRequest, run_forecast
from app.services.progress_service import read_progress, reset_progress, write_progress
from app.services.spending_slowdown import build_spending_slowdown
from src.pipeline.data_contracts import dedupe_columns_keep_first, first_series, normalize_date_series, enforce_schema
from app.utils.web_helpers import (
    safe_number as _safe_number,
    json_ready_records as _json_ready_records,
    plot_card as _plot_card,
    json_clean as _json_clean,
    safe_divide as _safe_divide,
    diagnostic_frequency_label as _diagnostic_frequency_label,
    monthly_equivalent_factor as _monthly_equivalent_factor,
    format_history_window_label as _format_history_window_label,
    build_driver_impact_frame as _build_driver_impact_frame,
    redirect,
    _load_upload_meta,
    _load_mapping_meta,
    _run_forecast_job,
    _load_result_payload,
    _safe_run_settings,
    _build_run_summary,
    _build_run_history_df,
    _build_compare_visuals,
    _clone_run_artifacts,
    _forecast_settings_path,
    _write_visuals_snapshot,
    _load_visuals_snapshot,
    _load_forecast_settings,
    _auto_select,
    _dataset_file_options,
    _build_regressor_columns,
    _resolve_dataset,
    _friendly_build_model_dataframe,
    _detect_frequency,
    _upload_meta_path,
    _mapping_meta_path,
    _project_dir,
    _build_quality_visuals,
    _build_results_visuals,
    _protect_results_visuals,
)

class CompatJinja2Templates(Jinja2Templates):
    def TemplateResponse(self, *args, **kwargs):
        if args and isinstance(args[0], str):
            name = args[0]
            context = args[1] if len(args) > 1 else kwargs.pop('context', None)
            context = context or {}
            request = context.get('request') or kwargs.pop('request', None)
            if request is None:
                raise ValueError('TemplateResponse context must include request')
            return super().TemplateResponse(request, name, context, *args[2:], **kwargs)
        return super().TemplateResponse(*args, **kwargs)


router = APIRouter()
templates = CompatJinja2Templates(directory=str(TEMPLATES_DIR))

from threading import Thread


@router.get('/copyright', response_class=HTMLResponse)
def copyright_page(request: Request):
    return templates.TemplateResponse('copyright.html', {
        'request': request,
        'app_name': APP_NAME,
        'app_version': APP_VERSION,
        'owner_name': 'Luis Rodriguez',
        'current_year': 2026,
    })


def _project_run_not_found_response(request: Request, project_id: int | None = None) -> HTMLResponse:
    dashboard_href = '/'
    projects_href = '/' if project_id is None else f'/projects/{project_id}/edit'
    html_body = f"""
    <!DOCTYPE html>
    <html lang='en'>
    <head>
      <meta charset='utf-8'>
      <meta name='viewport' content='width=device-width, initial-scale=1'>
      <title>SignalForge</title>
      <style>
        body {{ font-family: Arial, sans-serif; background: #0f172a; color: #e5e7eb; margin: 0; }}
        .wrap {{ min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 24px; }}
        .card {{ max-width: 640px; width: 100%; background: #111827; border: 1px solid #334155; border-radius: 18px; padding: 28px; box-shadow: 0 20px 50px rgba(0,0,0,.35); }}
        h1 {{ margin: 0 0 12px; font-size: 28px; }}
        p {{ color: #cbd5e1; line-height: 1.6; margin: 0 0 18px; }}
        .actions {{ display: flex; gap: 12px; flex-wrap: wrap; }}
        a {{ text-decoration: none; }}
        .btn {{ display: inline-block; padding: 12px 18px; border-radius: 12px; font-weight: 600; }}
        .btn-primary {{ background: #38bdf8; color: #082f49; }}
        .btn-secondary {{ background: #1f2937; color: #e5e7eb; border: 1px solid #475569; }}
      </style>
    </head>
    <body>
      <div class='wrap'>
        <div class='card'>
          <h1>Project or run not found</h1>
          <p>The project or run you tried to open is missing, was deleted, or the link is no longer valid.</p>
          <div class='actions'>
            <a class='btn btn-primary' href='{dashboard_href}'>Back to Dashboard</a>
            <a class='btn btn-secondary' href='{projects_href}'>Project Settings</a>
          </div>
        </div>
      </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html_body, status_code=404)


def _manual_adjustments_path(project_id: int, run_id: int) -> Path:
    return _project_dir(project_id) / f'run_{run_id}_manual_adjustments.json'


def _load_manual_adjustments(project_id: int, run_id: int) -> dict:
    path = _manual_adjustments_path(project_id, run_id)
    if not path.exists():
        return {
            'excluded_ranges': [],
            'outages': [],
            'promos': [],
            'forced_holidays': [],
            'notes': '',
        }
    try:
        data = json.loads(path.read_text(encoding='utf-8'))
        if not isinstance(data, dict):
            raise ValueError('manual adjustments payload must be an object')
    except Exception:
        return {
            'excluded_ranges': [],
            'outages': [],
            'promos': [],
            'forced_holidays': [],
            'notes': '',
        }
    data.setdefault('excluded_ranges', [])
    data.setdefault('outages', [])
    data.setdefault('promos', [])
    data.setdefault('forced_holidays', [])
    data.setdefault('notes', '')
    return data


def _build_training_window_context(df: pd.DataFrame | None, forecast_settings: dict | None = None) -> dict:
    ctx = {
        'actual_data_start': None,
        'actual_data_end': None,
        'all_history_start': None,
        'all_history_end': None,
        'preset_ranges': {},
    }
    if df is None or 'ds' not in df.columns:
        return ctx
    ds = pd.to_datetime(df['ds'], errors='coerce').dropna()
    if ds.empty:
        return ctx
    actual_start = pd.Timestamp(ds.min()).normalize()
    actual_end = pd.Timestamp(ds.max()).normalize()
    ctx['actual_data_start'] = str(actual_start.date())
    ctx['actual_data_end'] = str(actual_end.date())
    ctx['all_history_start'] = str(actual_start.date())
    ctx['all_history_end'] = str(actual_end.date())
    for key, days in {'last_90_days': 90, 'last_180_days': 180, 'last_365_days': 365}.items():
        start = (actual_end - pd.Timedelta(days=days - 1)).normalize()
        if start < actual_start:
            start = actual_start
        ctx['preset_ranges'][key] = {
            'start': str(start.date()),
            'end': str(actual_end.date()),
            'days': days,
        }
    return ctx


def _parse_adjustment_rows(starts: list[str], ends: list[str], labels: list[str], item_type: str) -> list[dict]:
    rows = []
    for start, end, label in zip(starts or [], ends or [], labels or []):
        start = (start or '').strip()
        end = (end or '').strip()
        label = (label or '').strip()
        if not start and not end and not label:
            continue
        rows.append({'start': start or None, 'end': end or None, 'label': label or item_type.replace('_', ' ').title()})
    return rows


@router.post('/projects/{project_id}/manual-adjustments')
def save_manual_adjustments(
    project_id: int,
    run_id: int = Form(...),
    excluded_start: list[str] = Form(default=[]),
    excluded_end: list[str] = Form(default=[]),
    excluded_reason: list[str] = Form(default=[]),
    outage_start: list[str] = Form(default=[]),
    outage_end: list[str] = Form(default=[]),
    outage_label: list[str] = Form(default=[]),
    promo_start: list[str] = Form(default=[]),
    promo_end: list[str] = Form(default=[]),
    promo_label: list[str] = Form(default=[]),
    forced_holiday_date: list[str] = Form(default=[]),
    forced_holiday_label: list[str] = Form(default=[]),
    adjustment_notes: str = Form(''),
):
    payload = {
        'excluded_ranges': _parse_adjustment_rows(excluded_start, excluded_end, excluded_reason, 'excluded_range'),
        'outages': _parse_adjustment_rows(outage_start, outage_end, outage_label, 'outage'),
        'promos': _parse_adjustment_rows(promo_start, promo_end, promo_label, 'promo'),
        'forced_holidays': [
            {'date': (dt or '').strip() or None, 'label': (label or '').strip() or 'Forced Holiday'}
            for dt, label in zip(forced_holiday_date or [], forced_holiday_label or [])
            if (dt or '').strip() or (label or '').strip()
        ],
        'notes': (adjustment_notes or '').strip(),
    }
    _manual_adjustments_path(project_id, run_id).write_text(json.dumps(payload, indent=2), encoding='utf-8')
    return redirect(f'/projects/{project_id}/quality?run_id={run_id}')


@router.get('/', response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    projects = list_projects(db)
    runs = list_runs(db)
    latest_run_by_project = {}
    for run in runs:
        latest_run_by_project.setdefault(run.project_id, run.id)
    active_statuses = {'running', 'queued', 'training', 'in_progress'}
    completed_statuses = {'completed', 'success', 'finished', 'done'}
    dashboard = {
        'project_count': len(projects),
        'run_count': len(runs),
        'active_count': sum(1 for run in runs if (run.status or '').lower() in active_statuses),
        'completed_count': sum(1 for run in runs if (run.status or '').lower() in completed_statuses),
        'latest_project': projects[0] if projects else None,
        'latest_run': runs[0] if runs else None,
    }
    return templates.TemplateResponse('home.html', {
        'request': request,
        'projects': projects,
        'runs': runs[:10],
        'latest_run_by_project': latest_run_by_project,
        'dashboard': dashboard,
        'app_name': APP_NAME,
        'app_version': APP_VERSION,
    })


@router.get('/projects/new', response_class=HTMLResponse)
def new_project_page(request: Request):
    return templates.TemplateResponse('project_new.html', {
        'request': request,
        'scope_options': [
            ('company_total', 'Company total only'),
            ('affiliate', 'Forecast by Affiliate ID'),
            ('platform', 'Forecast by Platform'),
            ('custom', 'Forecast by a custom category column'),
        ],
        'form_data': {
            'name': '',
            'description': '',
            'frequency': 'D',
            'years_of_history': 5,
            'grouped_forecast': False,
            'forecast_scope': 'company_total',
        },
        'error_message': None,
    })


@router.post('/projects')
def create_project_action(
    request: Request,
    name: str = Form(...),
    description: str = Form(''),
    frequency: str = Form('D'),
    years_of_history: int = Form(5),
    grouped_forecast: bool = Form(False),
    forecast_scope: str = Form('company_total'),
    db: Session = Depends(get_db),
):
    form_data = {
        'name': name,
        'description': description,
        'frequency': frequency,
        'years_of_history': years_of_history,
        'grouped_forecast': bool(grouped_forecast or forecast_scope != 'company_total'),
        'forecast_scope': forecast_scope,
    }
    try:
        project = create_project(db, ProjectCreate(
            name=name,
            description=description or None,
            frequency=frequency,
            years_of_history=years_of_history,
            grouped_forecast=(grouped_forecast or forecast_scope != 'company_total'),
            forecast_scope=forecast_scope,
        ))
    except ValueError as exc:
        if str(exc) == 'duplicate_project_name':
            return templates.TemplateResponse('project_new.html', {
                'request': request,
                'scope_options': [
                    ('company_total', 'Company total only'),
                    ('affiliate', 'Forecast by Affiliate ID'),
                    ('platform', 'Forecast by Platform'),
                    ('custom', 'Forecast by a custom category column'),
                ],
                'form_data': form_data,
                'error_message': f'A project named “{name}” already exists. Choose a different project name or delete the older one first.',
            }, status_code=409)
        raise
    run = create_run(db, project.id, 'Initial Run')
    return redirect(f'/projects/{project.id}/upload?run_id={run.id}')






@router.post('/projects/{project_id}/delete')
def delete_project_action(project_id: int, db: Session = Depends(get_db)):
    project = get_project(db, project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')

    for folder in [UPLOADS_DIR / f'project_{project_id}', PROJECTS_DIR / f'project_{project_id}']:
        try:
            if folder.exists():
                shutil.rmtree(folder, ignore_errors=True)
        except Exception:
            pass

    delete_project(db, project)
    return redirect('/')


@router.get('/projects/{project_id}/edit', response_class=HTMLResponse)
def edit_project_page(project_id: int, request: Request, run_id: int | None = None, db: Session = Depends(get_db)):
    project = get_project(db, project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    return templates.TemplateResponse('project_edit.html', {
        'request': request,
        'project': project,
        'run_id': run_id,
        'scope_options': [
            ('company_total', 'Company total only'),
            ('affiliate', 'Forecast by Affiliate ID'),
            ('platform', 'Forecast by Platform'),
            ('custom', 'Forecast by a custom category column'),
        ],
    })


@router.post('/projects/{project_id}/edit')
def edit_project_action(
    project_id: int,
    name: str = Form(...),
    description: str = Form(''),
    frequency: str = Form('D'),
    years_of_history: int = Form(5),
    grouped_forecast: bool = Form(False),
    forecast_scope: str = Form('company_total'),
    return_run_id: int | None = Form(None),
    db: Session = Depends(get_db),
):
    project = get_project(db, project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    update_project(db, project, ProjectCreate(
        name=name,
        description=description or None,
        frequency=frequency,
        years_of_history=years_of_history,
        grouped_forecast=(grouped_forecast or forecast_scope != 'company_total'),
        forecast_scope=forecast_scope,
    ))
    runs = list_runs(db, project_id=project_id)
    valid_run_ids = {r.id for r in runs}
    if return_run_id and return_run_id in valid_run_ids:
        target_run = return_run_id
    else:
        target_run = runs[0].id if runs else create_run(db, project_id, 'Edited Run').id
    return redirect(f'/projects/{project_id}/upload?run_id={target_run}')

@router.get('/projects/{project_id}/upload', response_class=HTMLResponse)
def upload_page(project_id: int, request: Request, run_id: int, db: Session = Depends(get_db)):
    project = get_project(db, project_id)
    run = get_run(db, run_id)
    if not project or not run:
        return _project_run_not_found_response(request, project_id)
    upload_meta = _load_upload_meta(project_id, run_id)
    return templates.TemplateResponse('upload.html', {
        'request': request,
        'project': project,
        'run': run,
        'upload_meta': upload_meta,
        'forecast_scope': getattr(project, 'forecast_scope', 'company_total'),
    })


@router.post('/projects/{project_id}/upload')
async def upload_action(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    form = await request.form()
    run_id = int(form.get('run_id'))
    project = get_project(db, project_id)
    run = get_run(db, run_id)
    if not project or not run:
        return _project_run_not_found_response(request, project_id)

    # Load any existing meta so we can preserve previously uploaded files
    existing_meta = _load_upload_meta(project_id, run_id)
    files_meta: dict[str, dict] = dict(existing_meta.get('files', {}))

    # Fixed named files
    fixed_keys = {
        'revenue_file': 'revenue',
        'leads_file': 'leads',
        'cost_file': 'cost',
        'events_file': 'events',
        'cohort_revenue_file': 'cohort_revenue',
    }
    for form_key, store_key in fixed_keys.items():
        upload = form.get(form_key)
        if upload and getattr(upload, 'filename', '') and upload.filename:
            saved = save_upload(project_id, run_id, upload)
            files_meta[store_key] = {'path': str(saved), 'filename': upload.filename}

    # Dynamic custom_N_file / custom_N_label fields
    import re
    custom_indices = set()
    for key in form.keys():
        m = re.match(r'^(custom_\d+)_(?:file|label)$', key)
        if m:
            custom_indices.add(m.group(1))

    for sheet_key in sorted(custom_indices):
        label_val = form.get(f'{sheet_key}_label', '').strip()
        upload = form.get(f'{sheet_key}_file')
        has_new_file = upload and getattr(upload, 'filename', '') and upload.filename
        if has_new_file:
            saved = save_upload(project_id, run_id, upload)
            files_meta[sheet_key] = {
                'path': str(saved),
                'filename': upload.filename,
                'label': label_val or sheet_key.replace('custom_', 'Custom '),
            }
        elif sheet_key in files_meta and label_val:
            # Update label only (file unchanged)
            files_meta[sheet_key]['label'] = label_val
        elif sheet_key not in files_meta and not has_new_file:
            # Row added but no file selected — skip
            pass

    meta_path = _upload_meta_path(project_id, run_id)
    meta_path.write_text(json.dumps({'files': files_meta}, indent=2), encoding='utf-8')
    return redirect(f'/projects/{project_id}/map?run_id={run_id}')





@router.post('/projects/{project_id}/upload/delete-file')
async def delete_uploaded_file_action(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    form = await request.form()
    run_id = int(form.get('run_id') or 0)
    file_key = str(form.get('file_key') or '').strip()
    project = get_project(db, project_id)
    run = get_run(db, run_id)
    if not project or not run:
        return _project_run_not_found_response(request, project_id)
    upload_meta = _load_upload_meta(project_id, run_id)
    files_meta = dict(upload_meta.get('files', {}))
    removed = files_meta.pop(file_key, None)
    if removed:
        try:
            file_path = Path(str(removed.get('path', ''))).expanduser()
            if file_path.exists() and file_path.is_file():
                file_path.unlink()
        except Exception:
            pass
        _upload_meta_path(project_id, run_id).write_text(json.dumps({'files': files_meta}, indent=2), encoding='utf-8')
    return redirect(f'/projects/{project_id}/upload?run_id={run_id}')


@router.get('/projects/{project_id}/map', response_class=HTMLResponse)
def mapping_page(project_id: int, request: Request, run_id: int, db: Session = Depends(get_db)):
    upload_meta = _load_upload_meta(project_id, run_id)
    project = get_project(db, project_id)
    if not upload_meta:
        return redirect(f'/projects/{project_id}/upload?run_id={run_id}')
    existing_mapping = _load_mapping_meta(project_id, run_id)
    # Ensure optional nested mappings always exist so the template can safely
    # access leads/cost/events fields even before the user has saved anything.
    existing_mapping.setdefault('revenue', {})
    existing_mapping.setdefault('leads', {})
    existing_mapping.setdefault('cost', {})
    existing_mapping.setdefault('events', {})
    existing_mapping.setdefault('cohort_revenue', {})
    existing_mapping.setdefault('custom_regressors', {})
    existing_mapping.setdefault('custom_regressor_columns', [])

    file_options = _dataset_file_options(upload_meta)
    file_labels = {key: label for key, label in file_options}

    # Auto-map patterns: each section maps field → (include patterns, exclude patterns)
    # Only applied when no existing saved mapping and no query param override.
    AUTO_PATTERNS: dict[str, dict[str, tuple]] = {
        'revenue': {
            'date_column':   ([[' date'], ['date '], ['revenue date'], ['date revenue'], ['report date'], ['period']], [['id']]),
            'target_column': ([[' revenue'], ['total revenue'], ['revenue amount'], ['revenue sum'], ['gross revenue'],
                               ['net revenue'], ['daily revenue'], ['monthly revenue'], ['weekly revenue'],
                               ['amount'], ['total amount'], ['sales'], ['total sales']], [['date'], ['id'], ['lead'], ['cost'], ['spend']]),
            'platform_column': ([['platform'], ['source'], ['channel']], [['id'], ['date']]),
            'affiliate_id_column': ([['affiliate id'], ['aff id'], ['affiliate']], [['date'], ['campaign'], ['ad']]),
            'campaign_id_column': ([['campaign id'], ['campaign']], [['date'], ['affiliate'], ['ad']]),
            'ad_id_column': ([['ad id'], ['ads id'], ['creative id'], ['adset id'], ['ad_id'], ['ads_id']], [['date'], ['month'], ['day'], ['affiliate'], ['campaign'], ['lead']]),
            'profile_id_column': ([['profile id'], ['profile_id'], ['member id'], ['user id'], ['customer id'], ['account id']], [['date'], ['campaign'], ['ad'], ['affiliate'], ['platform']]),
        },
        'leads': {
            'date_column':  ([['lead day'], ['lead date'], [' date'], ['date '], ['report date'], ['period'], ['day']], [['id']]),
            'value_column': ([['leads'], [' leads'], ['lead count'], ['leads count'], ['total leads'], ['sum of leads'],
                              ['new leads'], ['lead volume'], ['num leads'], ['# leads']], [['date'], ['id'], ['cost']]),
            'platform_column': ([['platform'], ['source'], ['channel']], [['id'], ['date']]),
            'affiliate_id_column': ([['affiliate id'], ['aff id'], ['affiliate']], [['date'], ['campaign'], ['ad']]),
            'campaign_id_column': ([['campaign id'], ['campaign']], [['date'], ['affiliate'], ['ad']]),
            'ad_id_column': ([['ad id'], ['ads id'], ['creative id'], ['adset id'], ['ad_id'], ['ads_id']], [['date'], ['month'], ['day'], ['affiliate'], ['campaign'], ['lead']]),
        },
        'cost': {
            'date_column':  ([[' date'], ['date '], ['spend date'], ['cost date'], ['report date'], ['period']], [['id']]),
            'value_column': ([[' cost'], ['total cost'], ['spend'], ['total spend'], ['ad spend'],
                              ['marketing spend'], ['ad cost'], ['media cost'], ['budget spent']], [['date'], ['id'], ['lead'], ['revenue']]),
            'platform_column': ([['platform'], ['source'], ['channel']], [['id'], ['date']]),
            'affiliate_id_column': ([['affiliate id'], ['aff id'], ['affiliate']], [['date'], ['campaign'], ['ad']]),
            'campaign_id_column': ([['campaign id'], ['campaign']], [['date'], ['affiliate'], ['ad']]),
            'ad_id_column': ([['ad id'], ['ads id'], ['creative id'], ['adset id'], ['ad_id'], ['ads_id']], [['date'], ['month'], ['day'], ['affiliate'], ['campaign'], ['lead']]),
        },
        'events': {
            'date_column':       ([[' date'], ['event date'], ['date '], ['period']], [['id']]),
            'event_flag_column': ([[' event'], ['promo'], ['holiday'], ['promotion'], ['campaign'], ['flag']], [['date'], ['outage']]),
            'outage_flag_column': ([[' outage'], ['downtime'], ['incident'], ['maintenance']], [['date']]),
        },
        'cohort_revenue': {
            'lead_month_column': ([['lead month'], ['cohort month'], ['lead cohort']], [['transaction'], ['revenue']]),
            'transaction_month_column': ([['transaction month'], ['revenue month'], ['txn month']], [['lead']]),
            'value_column': ([[' revenue'], ['total revenue'], ['amount'], ['sales']], [['lead'], ['transaction'], ['date']]),
            'platform_column': ([['platform'], ['source'], ['channel']], [['id'], ['date']]),
            'affiliate_id_column': ([['affiliate id'], ['aff id'], ['affiliate']], [['date'], ['campaign'], ['ad']]),
            'campaign_id_column': ([['campaign id'], ['campaign']], [['date'], ['affiliate'], ['ad']]),
            'ad_id_column': ([['ad id'], ['ads id'], ['creative id'], ['adset id'], ['ad_id'], ['ads_id']], [['date'], ['month'], ['day'], ['affiliate'], ['campaign'], ['lead']]),
        },
    }

    ID_FIELDS = {'affiliate_id_column', 'campaign_id_column', 'ad_id_column', 'profile_id_column'}

    def _is_safe_auto_id_column(field: str, column_name: str) -> bool:
        low = str(column_name or '').strip().lower().replace('_', ' ').replace('-', ' ')
        if field not in ID_FIELDS:
            return True
        reject_tokens = {'date', 'day', 'month', 'week', 'year', 'period'}
        if any(token in low for token in reject_tokens):
            return False
        if field == 'affiliate_id_column':
            return ('affiliate' in low or 'aff ' in low or low.endswith(' aff') or 'aff id' in low or 'affiliate id' in low) and 'id' in low
        if field == 'campaign_id_column':
            return 'campaign' in low and 'id' in low
        if field == 'ad_id_column':
            return any(token in low for token in ['ad id', 'ads id', 'adset id', 'creative id', 'ad_id', 'ads_id'])
        if field == 'profile_id_column':
            return any(token in low for token in ['profile id', 'profile_id', 'member id', 'user id', 'customer id', 'account id'])
        return True

    def _likely_helper_value_column(columns: list[str], df: pd.DataFrame | None, section_name: str, date_column: str | None) -> str | None:
        section_patterns = {
            'leads': [['leads'], ['lead count'], ['total leads'], ['lead volume'], ['num leads']],
            'cost': [['cost'], ['spend'], ['total cost'], ['total spend'], ['ad spend']],
        }
        excluded = {date_column} if date_column else set()
        suggestion = _auto_select(columns, section_patterns.get(section_name, []), [['date'], ['day'], ['month'], ['id']])
        if suggestion and suggestion not in excluded:
            return suggestion
        if df is None:
            return None
        candidates: list[tuple[str, float]] = []
        for col in columns:
            if col in excluded:
                continue
            low = str(col).lower().replace('_', ' ').replace('-', ' ')
            if any(token in low for token in ['date', 'day', 'month', 'period', 'id', 'platform', 'source', 'channel', 'affiliate', 'campaign', 'ad ']):
                continue
            series = pd.to_numeric(first_series(df, col), errors='coerce')
            valid_ratio = float(series.notna().mean()) if len(series) else 0.0
            median_val = float(series.dropna().median()) if series.notna().any() else 0.0
            if valid_ratio >= 0.75 and 0 <= median_val < 1e12:
                candidates.append((col, valid_ratio * 1000 - abs(median_val)))
        return sorted(candidates, key=lambda item: item[1], reverse=True)[0][0] if candidates else None

    def cfg_for(name: str, default_file: str | None = None) -> dict:
        cfg = existing_mapping.get(name, {}).copy()
        file_key = request.query_params.get(f'{name}_file_key') or cfg.get('file_key') or default_file
        sheet_name = request.query_params.get(f'{name}_sheet_name') or cfg.get('sheet_name')
        df, sheet_names, chosen_sheet = _resolve_dataset(upload_meta, file_key, sheet_name)
        columns = list(df.columns) if df is not None else []
        preview = df.head(8).fillna('').to_dict(orient='records') if df is not None else []
        # Restore any mapped column selections from query params (set by the JS reload)
        for field in ['date_column', 'value_column', 'category_column', 'value_columns',
                      'event_flag_column', 'outage_flag_column', 'lead_month_column', 'transaction_month_column',
                      'platform_column', 'affiliate_id_column', 'campaign_id_column', 'ad_id_column', 'profile_id_column', 'target_column']:
            qp = request.query_params.get(f'{name}_{field}')
            if qp is not None:
                cfg[field] = qp
        # Auto-select columns when nothing is saved yet for this section
        if columns and name in AUTO_PATTERNS:
            for field, (inc, exc) in AUTO_PATTERNS[name].items():
                if not cfg.get(field):
                    suggestion = _auto_select(columns, inc, exc)
                    if suggestion and _is_safe_auto_id_column(field, suggestion):
                        cfg[field] = suggestion
        if name in {'leads', 'cost'} and cfg.get('date_column') and cfg.get('value_column') == cfg.get('date_column'):
            replacement = _likely_helper_value_column(columns, df, name, cfg.get('date_column'))
            if replacement:
                cfg['value_column'] = replacement
        cfg.update({
            'file_key': file_key,
            'sheet_name': chosen_sheet,
            'sheet_names': sheet_names,
            'columns': columns,
            'preview': preview,
            'file_label': file_labels.get(file_key, file_key or 'None'),
        })
        return cfg

    revenue_cfg = cfg_for('revenue', 'revenue')
    leads_cfg = cfg_for('leads', 'leads')
    cost_cfg = cfg_for('cost', 'cost')
    events_cfg = cfg_for('events', 'events')
    cohort_revenue_cfg = cfg_for('cohort_revenue', 'cohort_revenue')
    custom_regressors_cfg = cfg_for('custom_regressors')

    auto_bound_sections = []
    for section_name, cfg, preferred_key in [
        ('Revenue', revenue_cfg, 'revenue'),
        ('Leads', leads_cfg, 'leads'),
        ('Cost / Spend', cost_cfg, 'cost'),
        ('Events', events_cfg, 'events'),
        ('Cohort Revenue', cohort_revenue_cfg, 'cohort_revenue'),
    ]:
        if cfg.get('file_key') == preferred_key and preferred_key in (upload_meta.get('files', {}) or {}):
            auto_bound_sections.append({
                'section': section_name,
                'filename': (upload_meta.get('files', {}).get(preferred_key) or {}).get('filename', preferred_key),
            })

    # Build a config dict for every custom_N sheet that was uploaded
    import re as _re
    custom_sheet_cfgs = {}
    for file_key in sorted(upload_meta.get('files', {}).keys()):
        if _re.match(r'^custom_\d+$', file_key):
            meta = upload_meta['files'][file_key]
            cfg = cfg_for(file_key, file_key)
            cfg['label'] = meta.get('label', file_key.replace('custom_', 'Custom '))
            existing_mapping.setdefault(file_key, {})
            custom_sheet_cfgs[file_key] = cfg

    return templates.TemplateResponse('mapping.html', {
        'request': request,
        'project_id': project_id,
        'run_id': run_id,
        'forecast_scope': getattr(project, 'forecast_scope', 'company_total') if project else 'company_total',
        'uploaded_files': upload_meta.get('files', {}),
        'file_options': file_options,
        'mapping': existing_mapping,
        'revenue': revenue_cfg,
        'leads': leads_cfg,
        'cost': cost_cfg,
        'events': events_cfg,
        'cohort_revenue': cohort_revenue_cfg,
        'custom_regressors': custom_regressors_cfg,
        'custom_sheet_cfgs': custom_sheet_cfgs,
        'auto_bound_sections': auto_bound_sections,
    })


@router.post('/projects/{project_id}/map')
async def mapping_action(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    import re as _re
    form = await request.form()

    def _f(key: str, default: str = '') -> str:
        return str(form.get(key, default) or default).strip()

    run_id = int(_f('run_id', '0'))
    revenue_date_column = _f('revenue_date_column')
    target_column = _f('target_column')
    category_column = _f('category_column')
    affiliate_id_column = _f('affiliate_id_column')
    platform_column = _f('platform_column')
    campaign_id_column = _f('campaign_id_column')
    ad_id_column = _f('ad_id_column')
    profile_id_column = _f('profile_id_column')
    custom_regressor_columns_raw = _f('custom_regressor_columns')
    leads_file_key = _f('leads_file_key')
    leads_sheet_name = _f('leads_sheet_name')
    leads_date_column = _f('leads_date_column')
    leads_column = _f('leads_column')
    leads_category_column = _f('leads_category_column')
    leads_platform_column = _f('leads_platform_column')
    leads_affiliate_id_column = _f('leads_affiliate_id_column')
    leads_campaign_id_column = _f('leads_campaign_id_column')
    leads_ad_id_column = _f('leads_ad_id_column')
    cost_file_key = _f('cost_file_key')
    cost_sheet_name = _f('cost_sheet_name')
    cost_date_column = _f('cost_date_column')
    cost_column = _f('cost_column')
    cost_category_column = _f('cost_category_column')
    cost_platform_column = _f('cost_platform_column')
    cost_affiliate_id_column = _f('cost_affiliate_id_column')
    cost_campaign_id_column = _f('cost_campaign_id_column')
    cost_ad_id_column = _f('cost_ad_id_column')
    events_file_key = _f('events_file_key')
    events_sheet_name = _f('events_sheet_name')
    events_date_column = _f('events_date_column')
    event_flag_column = _f('event_flag_column')
    outage_flag_column = _f('outage_flag_column')
    events_category_column = _f('events_category_column')
    cohort_revenue_file_key = _f('cohort_revenue_file_key')
    cohort_revenue_sheet_name = _f('cohort_revenue_sheet_name')
    cohort_lead_month_column = _f('cohort_revenue_lead_month_column')
    cohort_transaction_month_column = _f('cohort_revenue_transaction_month_column')
    cohort_revenue_value_column = _f('cohort_revenue_value_column')
    cohort_revenue_category_column = _f('cohort_revenue_category_column')
    cohort_revenue_platform_column = _f('cohort_revenue_platform_column')
    cohort_revenue_affiliate_id_column = _f('cohort_revenue_affiliate_id_column')
    cohort_revenue_campaign_id_column = _f('cohort_revenue_campaign_id_column')
    cohort_revenue_ad_id_column = _f('cohort_revenue_ad_id_column')
    custom_regressors_file_key = _f('custom_regressors_file_key')
    custom_regressors_sheet_name = _f('custom_regressors_sheet_name')
    custom_regressors_date_column = _f('custom_regressors_date_column')
    custom_regressors_value_columns = _f('custom_regressors_value_columns')
    custom_regressors_category_column = _f('custom_regressors_category_column')

    custom_regressors_list = [c.strip() for c in custom_regressor_columns_raw.split(',') if c.strip()]
    cr_value_cols = [c.strip() for c in custom_regressors_value_columns.split(',') if c.strip()]

    # Collect dynamic custom_N sheet mappings
    custom_sheet_keys = set()
    for key in form.keys():
        m = _re.match(r'^(custom_\d+)_(?:date_column|value_columns|category_column|file_key|sheet_name)$', key)
        if m:
            custom_sheet_keys.add(m.group(1))

    custom_sheets_mapping: dict[str, dict] = {}
    all_custom_regressor_cols = list(custom_regressors_list)
    for sheet_key in sorted(custom_sheet_keys):
        vc_raw = _f(f'{sheet_key}_value_columns')
        vc = [c.strip() for c in vc_raw.split(',') if c.strip()]
        custom_sheets_mapping[sheet_key] = {
            'file_key': _f(f'{sheet_key}_file_key') or sheet_key,
            'sheet_name': _f(f'{sheet_key}_sheet_name') or None,
            'date_column': _f(f'{sheet_key}_date_column') or None,
            'value_columns': vc,
            'category_column': _f(f'{sheet_key}_category_column') or None,
            'helper_frequency': _f(f'{sheet_key}_helper_frequency', 'auto'),
            'fill_method': _f(f'{sheet_key}_fill_method', 'distribute'),
        }
        # Merge value columns into global custom regressor list
        for col in vc:
            if col not in all_custom_regressor_cols:
                all_custom_regressor_cols.append(col)

    # Also merge from the dedicated custom_regressors section
    for col in cr_value_cols:
        if col not in all_custom_regressor_cols:
            all_custom_regressor_cols.append(col)

    mapping_payload = {
        'date_column': revenue_date_column,
        'target_column': target_column,
        'category_column': category_column or None,
        'affiliate_id_column': affiliate_id_column or None,
        'platform_column': platform_column or None,
        'campaign_id_column': campaign_id_column or None,
        'ad_id_column': ad_id_column or None,
        'profile_id_column': profile_id_column or None,
        'custom_regressor_columns': all_custom_regressor_cols,
        'revenue': {
            'file_key': _f('revenue_file_key', 'revenue'),
            'sheet_name': _f('revenue_sheet_name') or None,
            'date_column': revenue_date_column,
            'target_column': target_column,
            'category_column': category_column or None,
            'affiliate_id_column': affiliate_id_column or None,
            'platform_column': platform_column or None,
            'campaign_id_column': campaign_id_column or None,
            'ad_id_column': ad_id_column or None,
            'profile_id_column': profile_id_column or None,
        },
        'leads': {
            'file_key': leads_file_key or None,
            'sheet_name': leads_sheet_name or None,
            'date_column': leads_date_column or None,
            'value_column': leads_column or None,
            'category_column': (leads_platform_column or leads_category_column) or None,
            'platform_column': (leads_platform_column or leads_category_column) or None,
            'affiliate_id_column': leads_affiliate_id_column or None,
            'campaign_id_column': leads_campaign_id_column or None,
            'ad_id_column': leads_ad_id_column or None,
            'helper_frequency': _f('leads_helper_frequency', 'auto'),
            'fill_method': _f('leads_fill_method', 'distribute'),
        },
        'cost': {
            'file_key': cost_file_key or None,
            'sheet_name': cost_sheet_name or None,
            'date_column': cost_date_column or None,
            'value_column': cost_column or None,
            'category_column': (cost_platform_column or cost_category_column) or None,
            'platform_column': (cost_platform_column or cost_category_column) or None,
            'affiliate_id_column': cost_affiliate_id_column or None,
            'campaign_id_column': cost_campaign_id_column or None,
            'ad_id_column': cost_ad_id_column or None,
            'helper_frequency': _f('cost_helper_frequency', 'auto'),
            'fill_method': _f('cost_fill_method', 'distribute'),
        },
        'events': {
            'file_key': events_file_key or None,
            'sheet_name': events_sheet_name or None,
            'date_column': events_date_column or None,
            'event_flag_column': event_flag_column or None,
            'outage_flag_column': outage_flag_column or None,
            'category_column': events_category_column or None,
        },
        'cohort_revenue': {
            'file_key': cohort_revenue_file_key or None,
            'sheet_name': cohort_revenue_sheet_name or None,
            'lead_month_column': cohort_lead_month_column or None,
            'transaction_month_column': cohort_transaction_month_column or None,
            'value_column': cohort_revenue_value_column or None,
            'category_column': (cohort_revenue_platform_column or cohort_revenue_category_column) or None,
            'platform_column': (cohort_revenue_platform_column or cohort_revenue_category_column) or None,
            'affiliate_id_column': cohort_revenue_affiliate_id_column or None,
            'campaign_id_column': cohort_revenue_campaign_id_column or None,
            'ad_id_column': cohort_revenue_ad_id_column or None,
        },
        'leads_column': leads_column or None,
        'cost_column': cost_column or None,
        'event_flag_column': event_flag_column or None,
        'outage_flag_column': outage_flag_column or None,
        'custom_regressors': {
            'file_key': custom_regressors_file_key or None,
            'sheet_name': custom_regressors_sheet_name or None,
            'date_column': custom_regressors_date_column or None,
            'value_columns': cr_value_cols,
            'category_column': custom_regressors_category_column or None,
            'helper_frequency': _f('custom_regressors_helper_frequency', 'auto'),
            'fill_method': _f('custom_regressors_fill_method', 'distribute'),
        },
        **custom_sheets_mapping,
    }
    meta_path = _mapping_meta_path(project_id, run_id)
    meta_path.write_text(json.dumps(mapping_payload, indent=2), encoding='utf-8')
    return redirect(f'/projects/{project_id}/quality?run_id={run_id}')


@router.get('/projects/{project_id}/data-model', response_class=HTMLResponse)
def data_model_viewer(project_id: int, request: Request, run_id: int, db: Session = Depends(get_db)):
    upload_meta = _load_upload_meta(project_id, run_id)
    project = get_project(db, project_id)
    if not upload_meta:
        return redirect(f'/projects/{project_id}/upload?run_id={run_id}')
    mapping = _load_mapping_meta(project_id, run_id) or {}
    files = upload_meta.get('files', {}) or {}

    mapping.setdefault('revenue', {})
    mapping.setdefault('leads', {})
    mapping.setdefault('cost', {})
    mapping.setdefault('events', {})
    mapping.setdefault('cohort_revenue', {})
    mapping.setdefault('custom_regressors', {})

    def _clean_list(values):
        return [str(v).strip() for v in (values or []) if str(v).strip()]

    def _node_for(section_key: str, label: str, cfg: dict, fallback_file_key: str | None = None, kind: str = 'helper') -> dict | None:
        file_key = cfg.get('file_key') or fallback_file_key
        if not file_key or file_key not in files:
            return None
        df, sheet_names, chosen_sheet = _resolve_dataset(upload_meta, file_key, cfg.get('sheet_name'))
        columns = list(df.columns) if df is not None else []
        mapped_fields: list[dict] = []
        field_map = [
            ('date_column', 'Date'),
            ('value_column', 'Value'),
            ('target_column', 'Target'),
            ('platform_column', 'Platform'),
            ('affiliate_id_column', 'Affiliate'),
            ('campaign_id_column', 'Campaign'),
            ('ad_id_column', 'Ad'),
            ('profile_id_column', 'Profile'),
            ('event_flag_column', 'Event Flag'),
            ('outage_flag_column', 'Outage Flag'),
            ('lead_month_column', 'Lead Month'),
            ('transaction_month_column', 'Transaction Month'),
            ('category_column', 'Category'),
        ]
        for key, display in field_map:
            value = cfg.get(key)
            if value:
                mapped_fields.append({'label': display, 'value': value})
        value_columns = _clean_list(cfg.get('value_columns'))
        if value_columns:
            mapped_fields.append({'label': 'Regressors', 'value': ', '.join(value_columns[:4]) + (' +' if len(value_columns) > 4 else '')})
        helper_frequency = cfg.get('helper_frequency')
        fill_method = cfg.get('fill_method')
        return {
            'id': f'node-{section_key}',
            'section_key': section_key,
            'label': label,
            'kind': kind,
            'file_key': file_key,
            'file_label': files.get(file_key, {}).get('label') or files.get(file_key, {}).get('filename') or file_key,
            'filename': files.get(file_key, {}).get('filename') or file_key,
            'sheet_name': chosen_sheet,
            'sheet_names': sheet_names,
            'row_count': int(len(df)) if df is not None else 0,
            'column_count': len(columns),
            'columns_preview': columns[:8],
            'mapped_fields': mapped_fields,
            'helper_frequency': helper_frequency,
            'fill_method': fill_method,
            'is_primary': section_key == 'revenue',
        }

    revenue_cfg = {
        'file_key': (mapping.get('revenue') or {}).get('file_key') or 'revenue',
        'sheet_name': (mapping.get('revenue') or {}).get('sheet_name'),
        'date_column': mapping.get('date_column'),
        'target_column': mapping.get('target_column'),
        'category_column': mapping.get('category_column'),
        'platform_column': mapping.get('platform_column'),
        'affiliate_id_column': mapping.get('affiliate_id_column'),
        'campaign_id_column': mapping.get('campaign_id_column'),
        'ad_id_column': mapping.get('ad_id_column'),
        'profile_id_column': mapping.get('profile_id_column'),
        'value_columns': mapping.get('custom_regressor_columns', []),
    }
    dataset_nodes: list[dict] = []
    revenue_node = _node_for('revenue', 'Revenue', revenue_cfg, fallback_file_key='revenue', kind='primary')
    if revenue_node:
        dataset_nodes.append(revenue_node)

    helper_specs = [
        ('leads', 'Leads', mapping.get('leads') or {}, 'leads', 'helper'),
        ('cost', 'Cost / Spend', mapping.get('cost') or {}, 'cost', 'helper'),
        ('events', 'Events / Flags', mapping.get('events') or {}, 'events', 'helper'),
        ('cohort_revenue', 'Cohort Revenue', mapping.get('cohort_revenue') or {}, 'cohort_revenue', 'cohort'),
        ('custom_regressors', 'Custom Regressors', mapping.get('custom_regressors') or {}, None, 'custom'),
    ]
    for key, label, cfg, fallback, kind in helper_specs:
        node = _node_for(key, label, cfg, fallback_file_key=fallback, kind=kind)
        if node:
            dataset_nodes.append(node)
    for key in sorted(files.keys()):
        if not re.match(r'^custom_\d+$', key):
            continue
        cfg = mapping.get(key) or {}
        label = files.get(key, {}).get('label') or key.replace('custom_', 'Custom Sheet ')
        node = _node_for(key, label, cfg, fallback_file_key=key, kind='custom')
        if node:
            dataset_nodes.append(node)

    relationship_dimensions = []
    for field_key, label in [
        ('platform_column', 'Platform'),
        ('affiliate_id_column', 'Affiliate'),
        ('campaign_id_column', 'Campaign'),
        ('ad_id_column', 'Ad'),
        ('profile_id_column', 'Profile'),
    ]:
        if mapping.get(field_key):
            relationship_dimensions.append(label)

    relationships = []
    for node in dataset_nodes:
        if node['section_key'] == 'revenue':
            continue
        join_keys = ['Date']
        join_keys.extend(relationship_dimensions[:])
        if node['section_key'] in {'events', 'cohort_revenue'} and 'Category' not in join_keys:
            join_keys.append('Category')
        relation_note = 'Joined into the revenue timeline'
        if node.get('helper_frequency') and node['section_key'] in {'leads', 'cost', 'custom_regressors'} or node['section_key'].startswith('custom_'):
            relation_note = f"Expanded from {str(node.get('helper_frequency') or 'auto').upper()} using {str(node.get('fill_method') or 'distribute').replace('_', ' ')} before joining."
        relationships.append({
            'source_id': node['id'],
            'target_id': 'association-hub',
            'join_keys': join_keys,
            'note': relation_note,
        })

    model_df, build_error = _friendly_build_model_dataframe(upload_meta, mapping) if mapping else (None, None)
    regressor_columns = _build_regressor_columns(mapping)
    model_node = {
        'id': 'model-table',
        'label': 'Final Modeling Table',
        'row_count': int(len(model_df)) if model_df is not None else 0,
        'column_count': int(len(model_df.columns)) if model_df is not None else 0,
        'columns_preview': list(model_df.columns[:10]) if model_df is not None else [],
        'regressor_columns': regressor_columns,
        'date_column': mapping.get('date_column') or 'ds',
        'target_column': mapping.get('target_column') or 'y',
    }
    association_hub = {
        'label': 'Association Hub',
        'summary': 'SignalForge aligns helper datasets onto the revenue timeline, preserves mapped dimensions, and then builds one modeling table for training.',
        'keys': ['Date'] + relationship_dimensions if relationship_dimensions else ['Date'],
    }
    result_payload = _load_result_payload(project_id, run_id) or {}
    has_results = bool(result_payload)
    return templates.TemplateResponse('data_model.html', {
        'request': request,
        'project_id': project_id,
        'run_id': run_id,
        'project': project,
        'dataset_nodes': dataset_nodes,
        'relationships': relationships,
        'association_hub': association_hub,
        'model_node': model_node,
        'build_error': build_error,
        'has_results': has_results,
    })

@router.get('/projects/{project_id}/quality', response_class=HTMLResponse)
def quality_page(project_id: int, request: Request, run_id: int):
    import re as _re
    upload_meta = _load_upload_meta(project_id, run_id)
    mapping = _load_mapping_meta(project_id, run_id)
    if not mapping:
        return redirect(f'/projects/{project_id}/upload?run_id={run_id}')
    project = None
    try:
        from app.db import SessionLocal
        db = SessionLocal()
        try:
            project = get_project(db, project_id)
        finally:
            db.close()
    except Exception:
        project = None
    saved_settings = _load_forecast_settings(project_id, run_id) or {}
    default_frequency = getattr(project, 'frequency', None) or 'D'
    draft_forecast_settings = {
        'frequency': saved_settings.get('frequency', default_frequency),
        'horizon': int(saved_settings.get('horizon', 30) or 30),
        'scoring_metric': saved_settings.get('scoring_metric', 'balanced'),
        'output_mode': saved_settings.get('output_mode', 'scenario_bands'),
        'use_gpu': bool(saved_settings.get('use_gpu', False)),
        'enable_revenue_lag_modeling': bool(saved_settings.get('enable_revenue_lag_modeling', False)),
        'revenue_lag_profile': saved_settings.get('revenue_lag_profile', 'standard'),
        'business_profile': saved_settings.get('business_profile', 'general'),
        'training_window_mode': (saved_settings.get('training_window_mode') or 'all_history'),
        'training_window_start': (saved_settings.get('training_window_start') or '').strip() or None,
        'training_window_end': (saved_settings.get('training_window_end') or '').strip() or None,
        'models': [m for m in (saved_settings.get('models') or []) if m],
        'feature_engineering': saved_settings.get('feature_engineering', 'advanced_v2'),
    }

    auto_fix_summary = run_auto_fix_engine(project_id, run_id, upload_meta, mapping) if upload_meta else load_auto_fix_audit(project_id, run_id)
    upload_meta = _load_upload_meta(project_id, run_id)
    df, build_error = _friendly_build_model_dataframe(upload_meta, mapping)
    report = run_quality_checks(df, mapping['date_column'], mapping['target_column']) if df is not None else {}
    visuals = _build_quality_visuals(df)

    # Detect actual frequency for each helper so quality page shows the real resolved value
    def _resolved_freq(cfg: dict, file_key_default: str | None = None) -> str:
        fk = cfg.get('file_key') or file_key_default
        configured = cfg.get('helper_frequency', 'auto')
        if configured != 'auto':
            return configured
        # Auto — run detection on the actual file
        try:
            hdf, _, _ = _resolve_dataset(upload_meta, fk, cfg.get('sheet_name'))
            if hdf is not None:
                date_col = cfg.get('date_column') or mapping.get('date_column', '')
                if date_col and date_col in hdf.columns:
                    return _detect_frequency(hdf[date_col])
        except Exception:
            pass
        return 'auto'

    revenue_freq = 'D'
    if df is not None and mapping.get('date_column') in (df.columns if df is not None else []):
        revenue_freq = _detect_frequency(df[mapping['date_column']])

    freq_labels = {'D': 'Daily', 'W': 'Weekly', 'M': 'Monthly', 'auto': 'Auto-detect'}
    fill_labels = {'distribute': 'Distribute evenly', 'avg_monthly': 'Avg Monthly intelligent daily expansion', 'ffill': 'Forward-fill', 'interpolate': 'Interpolate'}
    freq_order = {'D': 0, 'W': 1, 'M': 2}

    dataset_info = []
    for section_key in ['leads', 'cost', 'custom_regressors']:
        cfg = mapping.get(section_key, {})
        if not cfg or not cfg.get('file_key'):
            continue
        label = {'leads': 'Leads', 'cost': 'Cost / Spend', 'custom_regressors': 'Custom Regressors'}.get(section_key, section_key)
        resolved = _resolved_freq(cfg)
        fill = cfg.get('fill_method', 'distribute')
        value_cols = cfg.get('value_columns') or ([cfg.get('value_column')] if cfg.get('value_column') else [])
        resampled = freq_order.get(resolved, 0) > freq_order.get(revenue_freq, 0)
        dataset_info.append({
            'label': label,
            'resolved_freq': freq_labels.get(resolved, resolved),
            'fill_method': fill_labels.get(fill, fill),
            'value_cols': [c for c in value_cols if c],
            'resampled': resampled,
            'revenue_freq': freq_labels.get(revenue_freq, revenue_freq),
        })
    for key, val in mapping.items():
        if _re.match(r'^custom_\d+$', key) and isinstance(val, dict) and val.get('file_key'):
            files = upload_meta.get('files', {})
            label = (files.get(key) or {}).get('label', key.replace('custom_', 'Custom Sheet '))
            resolved = _resolved_freq(val, key)
            fill = val.get('fill_method', 'distribute')
            resampled = freq_order.get(resolved, 0) > freq_order.get(revenue_freq, 0)
            dataset_info.append({
                'label': label,
                'resolved_freq': freq_labels.get(resolved, resolved),
                'fill_method': fill_labels.get(fill, fill),
                'value_cols': [c for c in val.get('value_columns', []) if c],
                'resampled': resampled,
                'revenue_freq': freq_labels.get(revenue_freq, revenue_freq),
            })

    forecast_category = _resolve_forecast_category_column(project, mapping)
    series_audit = _build_merged_series_audit(df, forecast_category)
    manual_adjustments = _load_manual_adjustments(project_id, run_id)
    training_window_context = _build_training_window_context(df, saved_settings)

    connection_nodes = [
        {'label': 'Cost', 'enabled': bool(mapping.get('cost') and mapping.get('cost', {}).get('file_key')), 'detail': mapping.get('cost', {}).get('value_column') or 'Not mapped'},
        {'label': 'Leads', 'enabled': bool(mapping.get('leads') and mapping.get('leads', {}).get('file_key')), 'detail': mapping.get('leads', {}).get('value_column') or mapping.get('leads_column') or 'Not mapped'},
        {'label': 'Revenue', 'enabled': bool(mapping.get('target_column')), 'detail': mapping.get('target_column') or 'Not mapped'},
        {'label': 'Cohort Revenue', 'enabled': bool(mapping.get('cohort_revenue') and mapping.get('cohort_revenue', {}).get('file_key')), 'detail': (mapping.get('cohort_revenue') or {}).get('value_column') or 'Optional'},
        {'label': 'Events', 'enabled': bool(mapping.get('events') and mapping.get('events', {}).get('file_key')), 'detail': (mapping.get('events') or {}).get('value_column') or 'Optional'},
    ]
    quality_guidance = build_quality_guidance(mapping, dataset_info, training_window_context, manual_adjustments, saved_settings)

    connection_badges = [
        {'label': 'Date Link', 'value': mapping.get('date_column') or 'Not mapped'},
        {'label': 'Revenue Target', 'value': mapping.get('target_column') or 'Not mapped'},
        {'label': 'Platform Link', 'value': mapping.get('platform_column') or ((mapping.get('leads') or {}).get('platform_column')) or ((mapping.get('cost') or {}).get('platform_column')) or 'Not mapped'},
        {'label': 'Affiliate Link', 'value': mapping.get('affiliate_id_column') or 'Not mapped'},
        {'label': 'Campaign Link', 'value': mapping.get('campaign_id_column') or 'Not mapped'},
        {'label': 'Lag-Aware Pairing', 'value': 'Enabled' if saved_settings.get('enable_revenue_lag_modeling') else 'Off'},
    ]

    return templates.TemplateResponse('quality.html', {
        'request': request,
        'project_id': project_id,
        'run_id': run_id,
        'report': report,
        'auto_fix_summary': auto_fix_summary,
        'mapping': mapping,
        'uploaded_files': upload_meta.get('files', {}),
        'build_error': build_error,
        'visuals': visuals,
        'dataset_info': dataset_info,
        'series_audit': series_audit,
        'forecast_category': forecast_category,
        'manual_adjustments': manual_adjustments,
        'training_window_context': training_window_context,
        'connection_nodes': connection_nodes,
        'connection_badges': connection_badges,
        'quality_guidance': quality_guidance,
    })



def _resolve_forecast_category_column(project, mapping: dict) -> str | None:
    scope = (getattr(project, 'forecast_scope', None) or 'company_total') if project is not None else 'company_total'
    if scope == 'company_total':
        return None
    if scope == 'platform':
        return mapping.get('platform_column') or mapping.get('category_column')
    if scope == 'affiliate':
        return mapping.get('affiliate_id_column')
    if scope == 'custom':
        return mapping.get('category_column') or mapping.get('platform_column')
    return mapping.get('category_column')


def _build_merged_series_audit(df: pd.DataFrame | None, category_column: str | None = None) -> dict:
    audit = {
        'raw_rows': 0,
        'raw_distinct_dates': 0,
        'raw_duplicate_date_rows': 0,
        'raw_total_y': None,
        'final_rows': 0,
        'final_distinct_dates': 0,
        'final_duplicate_date_rows': 0,
        'final_total_y': None,
        'final_median_y': None,
        'final_max_y': None,
        'final_forecast_to_history_hint': None,
        'category_mode': category_column or 'company_total',
    }
    if df is None or df.empty or 'ds' not in df.columns or 'y' not in df.columns:
        return audit
    temp = df.copy()
    temp['ds'] = pd.to_datetime(first_series(temp, 'ds'), errors='coerce').dt.normalize()
    temp['y'] = pd.to_numeric(first_series(temp, 'y'), errors='coerce')
    temp = temp.dropna(subset=['ds'])
    if temp.empty:
        return audit
    audit['raw_rows'] = int(len(temp))
    audit['raw_distinct_dates'] = int(temp['ds'].nunique())
    audit['raw_duplicate_date_rows'] = int(temp.duplicated(subset=['ds']).sum())
    raw_total_y = float(temp['y'].sum()) if temp['y'].notna().any() else None
    audit['raw_total_y'] = raw_total_y
    agg = {'y': 'sum'}
    for col in ['leads', 'cost']:
        if col in temp.columns:
            agg[col] = 'sum'
    rolled = temp.groupby('ds', as_index=False).agg(agg).sort_values('ds')
    audit['final_rows'] = int(len(rolled))
    audit['final_distinct_dates'] = int(rolled['ds'].nunique())
    audit['final_duplicate_date_rows'] = int(rolled.duplicated(subset=['ds']).sum())
    final_total_y = float(rolled['y'].sum()) if rolled['y'].notna().any() else None
    audit['final_total_y'] = final_total_y
    audit['final_median_y'] = float(rolled['y'].median()) if rolled['y'].notna().any() else None
    audit['final_max_y'] = float(rolled['y'].max()) if rolled['y'].notna().any() else None
    recent = rolled['y'].tail(min(30, len(rolled))).dropna()
    hist = rolled['y'].dropna()
    if not recent.empty and not hist.empty and float(hist.mean()) != 0:
        audit['final_forecast_to_history_hint'] = round(float(recent.mean()) / float(hist.mean()), 4)
    return audit



@router.get('/projects/{project_id}/forecast', response_class=HTMLResponse)
def forecast_page(project_id: int, request: Request, run_id: int, db: Session = Depends(get_db)):
    mapping = _load_mapping_meta(project_id, run_id)
    upload_meta = _load_upload_meta(project_id, run_id)
    project = get_project(db, project_id)
    forecast_settings = _load_forecast_settings(project_id, run_id)
    model_df = None
    try:
        model_df, _ = _friendly_build_model_dataframe(upload_meta, mapping)
    except Exception:
        model_df = None
    training_window_context = _build_training_window_context(model_df, forecast_settings)
    forecast_guidance = build_forecast_guidance(mapping, training_window_context, forecast_settings)
    return templates.TemplateResponse('forecast_setup.html', {
        'request': request,
        'project_id': project_id,
        'run_id': run_id,
        'project': project,
        'mapping': mapping,
        'uploaded_files': upload_meta.get('files', {}),
        'forecast_settings': forecast_settings,
        'training_window_context': training_window_context,
        'forecast_guidance': forecast_guidance,
    })


@router.post('/projects/{project_id}/forecast')
def forecast_action(
    project_id: int,
    request: Request,
    run_id: int = Form(...),
    frequency: str = Form('D'),
    horizon: int = Form(30),
    scoring_metric: str = Form('balanced'),
    output_mode: str = Form('scenario_bands'),
    use_gpu: bool = Form(False),
    enable_revenue_lag_modeling: bool = Form(False),
    revenue_lag_profile: str = Form('standard'),
    business_profile: str = Form('general'),
    training_window_mode: str = Form('all_history'),
    training_window_start: str = Form(''),
    training_window_end: str = Form(''),
    models: list[str] = Form(default=[]),
    db: Session = Depends(get_db),
):
    upload_meta = _load_upload_meta(project_id, run_id)
    mapping = _load_mapping_meta(project_id, run_id)
    project = get_project(db, project_id)
    if not upload_meta:
        return redirect(f'/projects/{project_id}/upload?run_id={run_id}')
    df, build_error = _friendly_build_model_dataframe(upload_meta, mapping)
    if df is None:
        _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
        return redirect(f'/projects/{project_id}/map?run_id={run_id}')
    selected_models = [m for m in models if m]

    training_window_start = (training_window_start or '').strip() or None
    training_window_end = (training_window_end or '').strip() or None
    actual_dates = pd.to_datetime(df[mapping['date_column']], errors='coerce').dropna()
    actual_end_dt = pd.Timestamp(actual_dates.max()).normalize() if not actual_dates.empty else None
    actual_start_dt = pd.Timestamp(actual_dates.min()).normalize() if not actual_dates.empty else None
    if training_window_mode in {'last_90_days','last_180_days','last_365_days'} and actual_end_dt is not None:
        days = {'last_90_days':90,'last_180_days':180,'last_365_days':365}[training_window_mode]
        preset_start = (actual_end_dt - pd.Timedelta(days=days - 1)).normalize()
        if actual_start_dt is not None and preset_start < actual_start_dt:
            preset_start = actual_start_dt
        training_window_start = str(preset_start.date())
        training_window_end = str(actual_end_dt.date())
    elif training_window_mode == 'all_history' and actual_start_dt is not None and actual_end_dt is not None:
        training_window_start = None
        training_window_end = None
    if training_window_mode == 'custom_range' and (not training_window_start or not training_window_end):
        _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
        return redirect(f'/projects/{project_id}/forecast?run_id={run_id}&model_error=Select+a+training+start+and+end+date')
    if training_window_mode == 'custom_range' and training_window_start and training_window_end:
        try:
            start_dt = pd.to_datetime(training_window_start)
            end_dt = pd.to_datetime(training_window_end)
            if end_dt < start_dt:
                _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
                return redirect(f'/projects/{project_id}/forecast?run_id={run_id}&model_error=Training+end+date+must+be+after+start+date')
            filtered_check = df[(pd.to_datetime(df[mapping['date_column']], errors='coerce') >= start_dt) & (pd.to_datetime(df[mapping['date_column']], errors='coerce') <= end_dt)]
            if len(filtered_check) < 60:
                _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
                return redirect(f'/projects/{project_id}/forecast?run_id={run_id}&model_error=Training+window+is+too+short.+Use+at+least+60+rows')
        except Exception:
            _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
            return redirect(f'/projects/{project_id}/forecast?run_id={run_id}&model_error=Training+window+dates+could+not+be+parsed')
    elif training_window_mode in {'last_90_days','last_180_days','last_365_days'}:
        try:
            start_dt = pd.to_datetime(training_window_start, errors='coerce')
            end_dt = pd.to_datetime(training_window_end, errors='coerce')
            ds_series = pd.to_datetime(df[mapping['date_column']], errors='coerce')
            filtered_check = df[(ds_series >= start_dt) & (ds_series <= end_dt)]
            if len(filtered_check) < 60:
                _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
                return redirect(f'/projects/{project_id}/forecast?run_id={run_id}&model_error=Selected+training+window+is+too+short+for+stable+training')
        except Exception:
            pass
    if not selected_models:
        _forecast_settings_path(project_id, run_id).write_text(json.dumps(draft_forecast_settings, indent=2), encoding='utf-8')
        return redirect(f'/projects/{project_id}/forecast?run_id={run_id}&model_error=Select+at+least+one+model')
    forecast_settings = {
        'frequency': frequency,
        'horizon': horizon,
        'scoring_metric': scoring_metric,
        'output_mode': output_mode,
        'use_gpu': bool(use_gpu),
        'enable_revenue_lag_modeling': bool(enable_revenue_lag_modeling),
        'revenue_lag_profile': revenue_lag_profile,
        'business_profile': business_profile,
        'models': selected_models,
        'training_window_mode': training_window_mode,
        'training_window_start': training_window_start,
        'training_window_end': training_window_end,
        'feature_engineering': 'advanced_v2',
    }
    _forecast_settings_path(project_id, run_id).write_text(json.dumps(forecast_settings, indent=2), encoding='utf-8')
    payload = ForecastRequest(
        df=df,
        date_column=mapping['date_column'],
        target_column=mapping['target_column'],
        category_column=_resolve_forecast_category_column(project, mapping),
        frequency=frequency,
        horizon=horizon,
        scoring_metric=scoring_metric,
        selected_models=selected_models,
        output_mode=output_mode,
        use_gpu=use_gpu,
        enable_revenue_lag_modeling=enable_revenue_lag_modeling,
        revenue_lag_profile=revenue_lag_profile,
        business_profile=business_profile,
        training_window_mode=training_window_mode,
        training_window_start=training_window_start,
        training_window_end=training_window_end,
        regressor_columns=_build_regressor_columns(mapping),
    )
    reset_progress(project_id, run_id)
    Thread(target=_run_forecast_job, args=(project_id, run_id, payload), daemon=True).start()
    return redirect(f'/projects/{project_id}/running?run_id={run_id}')


@router.get('/projects/{project_id}/running', response_class=HTMLResponse)
def running_page(project_id: int, request: Request, run_id: int):
    return templates.TemplateResponse('running.html', {
        'request': request,
        'project_id': project_id,
        'run_id': run_id,
    })


@router.get('/api/projects/{project_id}/progress')
def progress_api(project_id: int, run_id: int):
    return JSONResponse(read_progress(project_id, run_id))




@router.post('/projects/{project_id}/runs/{run_id}/rerun')
def rerun_from_saved_config(project_id: int, run_id: int, db: Session = Depends(get_db)):
    project = get_project(db, project_id)
    source_run = get_run(db, run_id)
    if not project or not source_run:
        return _project_run_not_found_response(request, project_id)
    upload_meta = _load_upload_meta(project_id, run_id)
    mapping = _load_mapping_meta(project_id, run_id)
    settings = _load_forecast_settings(project_id, run_id)
    if not upload_meta or not mapping or not settings:
        raise HTTPException(status_code=400, detail='This run does not have enough saved configuration to rerun yet.')
    new_run = create_run(db, project_id, f'Rerun of {source_run.label}')
    _clone_run_artifacts(project_id, run_id, new_run.id)
    df, build_error = _friendly_build_model_dataframe(upload_meta, mapping)
    if df is None:
        raise HTTPException(status_code=400, detail=f'Unable to rebuild the modeling table. {build_error}')
    payload = ForecastRequest(
        df=df,
        date_column=mapping['date_column'],
        target_column=mapping['target_column'],
        category_column=_resolve_forecast_category_column(project, mapping),
        frequency=settings.get('frequency', project.frequency),
        horizon=int(settings.get('horizon', 30)),
        scoring_metric=settings.get('scoring_metric', 'balanced'),
        selected_models=settings.get('models') or ['naive', 'seasonal_naive', 'autoarima', 'autoets', 'prophet', 'lightgbm', 'xgboost', 'random_forest'],
        output_mode=settings.get('output_mode', 'scenario_bands'),
        use_gpu=bool(settings.get('use_gpu', False)),
        enable_revenue_lag_modeling=bool(settings.get('enable_revenue_lag_modeling', False)),
        revenue_lag_profile=settings.get('revenue_lag_profile', 'standard'),
        business_profile=settings.get('business_profile', 'general'),
        training_window_mode=settings.get('training_window_mode', 'all_history'),
        training_window_start=settings.get('training_window_start') or None,
        training_window_end=settings.get('training_window_end') or None,
        regressor_columns=_build_regressor_columns(mapping),
    )
    reset_progress(project_id, new_run.id)
    Thread(target=_run_forecast_job, args=(project_id, new_run.id, payload), daemon=True).start()
    return redirect(f'/projects/{project_id}/running?run_id={new_run.id}')


@router.post('/projects/{project_id}/runs/{run_id}/duplicate')
def duplicate_run_config(project_id: int, run_id: int, db: Session = Depends(get_db)):
    project = get_project(db, project_id)
    source_run = get_run(db, run_id)
    if not project or not source_run:
        return _project_run_not_found_response(request, project_id)
    new_run = create_run(db, project_id, f'Copy of {source_run.label}')
    _clone_run_artifacts(project_id, run_id, new_run.id)
    return redirect(f'/projects/{project_id}/forecast?run_id={new_run.id}')





def _results_visual_defaults() -> dict:
    return {
        'charts': [],
        'forecast_table': [],
        'overview_cards': {},
        'insights': {},
        'cohort_table': [],
        'series_table': [],
        'optimizer_table': [],
        'time_machine_table': [],
        'drivers_table': [],
        'tracking_table': [],
        'alerts': [],
        'impact_table': [],
        'allocation_table': [],
        'model_ranking': [],
        'diagnostics_rows': [],
        'sheet_impact_table': [],
        'cohort_rev_available': False,
        'cohort_rev_summary': {},
        'cohort_rev_table': [],
        'cohort_rev_platform_table': [],
        'cohort_rev_curve_chart': None,
        'cohort_rev_platform_chart': None,
        'cohort_rev_heatmap_chart': None,
        'cohort_rev_payback_chart': None,
        'cohort_rev_payback_table': [],
        'cohort_rev_narrative': [],
        'cohort_rev_roas_table': [],
        'cohort_payback_curve': [],
        'cohort_payback_summary': [],
        'early_warning_table': [],
        'waterfall_table': [],
        'goal_optimizer_profiles': [],
        'planning_dimension_recommendations': [],
        'planning_whale_watch': [],
        'planning_quarter_horizon_months': 5,
        'statistics_available': False,
        'statistics_cards': {},
        'statistics_charts': [],
        'statistics_lag_table': [],
        'statistics_frontier_table': [],
        'statistics_attrition_table': [],
        'cost_validation_available': False,
        'cost_validation_cards': {},
        'cost_validation_platform_table': [],
        'cost_validation_company_table': [],
        'scenario_compare_table': [],
        'scenario_compare_chart': None,
        'budget_opt_table': [],
        'budget_opt_chart': None,
        'anomaly_table': [],
        'anomaly_chart': None,
        'rolling_update_table': [],
        'rolling_update_chart': None,
        'health_monitor_table': [],
        'health_monitor_cards': {},
        'auto_insights_table': [],
        'spending_slowdown_available': False,
        'spending_slowdown_message': '',
        'spending_slowdown_cards': {},
        'spending_slowdown_charts': [],
        'spending_slowdown_decomposition_table': [],
        'spending_slowdown_returning_table': [],
        'spending_slowdown_cohort_table': [],
        'forecast_start': None,
        'render_warning': None,
        'render_error': None,
        'lightgbm_intelligence_available': False,
        'lightgbm_exec_summary': [],
        'lightgbm_shap_summary': [],
        'lightgbm_local_explanations': [],
        'lightgbm_shap_aggregations': [],
        'lightgbm_ranked_opportunities': [],
        'lightgbm_lead_quality_scores': [],
        'lightgbm_anomaly_rows': [],
        'lightgbm_whale_predictions': [],
        'lightgbm_response_curve': [],
        'lightgbm_scenarios': [],
        'lightgbm_affiliate_quality': [],
        'lightgbm_confidence_rows': [],
    }


def _merge_results_visual_defaults(visuals: dict | None) -> dict:
    merged = _results_visual_defaults()
    if isinstance(visuals, dict):
        merged.update(visuals)
    return merged



def _snapshot_has_planning_payload(visuals: dict | None) -> bool:
    if not isinstance(visuals, dict):
        return False
    try:
        insights = visuals.get('insights') if isinstance(visuals.get('insights'), dict) else {}
        scenario_defaults = insights.get('scenario_defaults') if isinstance(insights, dict) else {}
        goal_profiles = visuals.get('goal_optimizer_profiles')
        has_defaults = isinstance(scenario_defaults, dict) and any(v not in (None, '', 0, 0.0, [], {}) for v in scenario_defaults.values())
        has_profiles = isinstance(goal_profiles, list) and len(goal_profiles) > 0
        return bool(has_defaults or has_profiles)
    except Exception:
        return False
def _snapshot_has_renderable_charts(visuals: dict | None) -> bool:
    if not isinstance(visuals, dict):
        return False

    def _chart_has_spec(chart: dict | None) -> bool:
        if not isinstance(chart, dict):
            return False
        spec = chart.get('spec')
        if isinstance(spec, str):
            return bool(spec.strip())
        return isinstance(spec, dict) and bool(spec)

    for key in ('charts', 'statistics_charts', 'spending_slowdown_charts'):
        items = visuals.get(key)
        if isinstance(items, list) and any(_chart_has_spec(item) for item in items):
            return True

    for key in ('anomaly_chart', 'rolling_update_chart', 'scenario_compare_chart', 'budget_opt_chart'):
        if _chart_has_spec(visuals.get(key)):
            return True

    registry = visuals.get('chart_registry')
    if isinstance(registry, dict):
        for spec in registry.values():
            if isinstance(spec, str) and spec.strip():
                return True
            if isinstance(spec, dict) and spec:
                return True
    return False


def _normalize_company_total_payback_labels(visuals: dict | None) -> dict | None:
    if not isinstance(visuals, dict):
        return visuals
    try:
        chart = visuals.get('cohort_rev_payback_chart')
        traces = chart.get('data') if isinstance(chart, dict) else None
        if isinstance(traces, list) and traces:
            categories: list[str] = []
            parsed_names: list[tuple[dict, str, str]] = []
            for trace in traces:
                if not isinstance(trace, dict):
                    continue
                raw_name = str(trace.get('name') or '')
                marker = '·' if '·' in raw_name else ('Â·' if 'Â·' in raw_name else None)
                if not marker:
                    continue
                lead_month, category = raw_name.split(marker, 1)
                lead_month = lead_month.strip()
                category = category.strip()
                if not category:
                    continue
                categories.append(category)
                parsed_names.append((trace, lead_month, category))
            unique_categories = {c for c in categories if c}
            if len(unique_categories) == 1:
                only_category = next(iter(unique_categories))
                if only_category.startswith('Platform=') or only_category.startswith('Group='):
                    for trace, lead_month, _ in parsed_names:
                        trace['name'] = f'{lead_month} · Company total'

        payback_table = visuals.get('cohort_rev_payback_table')
        if isinstance(payback_table, list) and payback_table:
            categories = {
                str(row.get('cohort_category') or '').strip()
                for row in payback_table
                if isinstance(row, dict) and str(row.get('cohort_category') or '').strip()
            }
            if len(categories) == 1:
                only_category = next(iter(categories))
                if only_category.startswith('Platform=') or only_category.startswith('Group='):
                    for row in payback_table:
                        if isinstance(row, dict):
                            row['cohort_category'] = 'Company total'
    except Exception:
        return visuals
    return visuals


def _normalize_cohort_maturity_labels(visuals: dict | None) -> dict | None:
    if not isinstance(visuals, dict):
        return visuals
    try:
        def _normalize_roas_rows(rows: object) -> None:
            if not isinstance(rows, list):
                return
            for row in rows:
                if not isinstance(row, dict):
                    continue
                try:
                    roas_value = float(row.get('roas')) if row.get('roas') not in (None, '', 'None') else None
                except Exception:
                    roas_value = None
                if roas_value is not None:
                    row['roas_pct'] = roas_value * 100.0
                    if 'payback_pct' in row:
                        row['payback_pct'] = max(0.0, min(100.0, roas_value * 100.0))
                elif 'payback_pct' in row:
                    try:
                        existing_value = float(row.get('payback_pct'))
                    except Exception:
                        existing_value = None
                    if existing_value is not None:
                        row['payback_pct'] = max(0.0, min(100.0, existing_value))

        _normalize_roas_rows(visuals.get('cohort_rev_roas_table'))
        _normalize_roas_rows(visuals.get('cohort_payback_summary'))
        _normalize_roas_rows(visuals.get('cohort_rev_payback_curve'))

        rows = visuals.get('cohort_rev_roas_table')
        if not isinstance(rows, list):
            return visuals
        for row in rows:
            if not isinstance(row, dict):
                continue
            active_months = row.get('active_months')
            try:
                active_months_num = float(active_months)
            except Exception:
                continue
            if active_months_num >= 6:
                row['maturity_status'] = 'Mature'
            elif active_months_num >= 3:
                row['maturity_status'] = 'Growing'
            else:
                row['maturity_status'] = 'Early'
    except Exception:
        return visuals
    return visuals


def _build_minimal_results_visuals(result: dict) -> dict:
    forecast_rows = result.get('blended_forecast') or result.get('best_forecast') or []
    forecast_table: list[dict] = []
    chart_spec = None
    if forecast_rows:
        for row in forecast_rows[:500]:
            if isinstance(row, dict):
                forecast_table.append({k: row.get(k) for k in row.keys()})
        ds = [row.get('ds') for row in forecast_rows if isinstance(row, dict)]
        yhat = [row.get('yhat') for row in forecast_rows if isinstance(row, dict)]
        lower = [row.get('yhat_lower') for row in forecast_rows if isinstance(row, dict)]
        upper = [row.get('yhat_upper') for row in forecast_rows if isinstance(row, dict)]
        chart_spec = {
            'data': [
                {'type': 'scatter', 'mode': 'lines', 'name': 'Forecast', 'x': ds, 'y': yhat},
                {'type': 'scatter', 'mode': 'lines', 'name': 'Lower', 'x': ds, 'y': lower, 'line': {'dash': 'dot'}},
                {'type': 'scatter', 'mode': 'lines', 'name': 'Upper', 'x': ds, 'y': upper, 'line': {'dash': 'dot'}},
            ],
            'layout': {
                'margin': {'l': 48, 'r': 18, 't': 24, 'b': 48},
                'paper_bgcolor': 'rgba(0,0,0,0)',
                'plot_bgcolor': 'rgba(0,0,0,0)',
                'legend': {'orientation': 'h'},
                'xaxis': {'title': 'Date'},
                'yaxis': {'title': 'Forecast'},
            },
        }
    model_ranking = result.get('top_models') or []
    diagnostics = result.get('diagnostics') or {}
    overview_cards = {
        'forecast_start': forecast_rows[0].get('ds') if forecast_rows and isinstance(forecast_rows[0], dict) else 'n/a',
        'latest_actual': diagnostics.get('recent_avg_y'),
        'forecast_avg': diagnostics.get('forecast_avg_expected'),
        'forecast_to_history_ratio': diagnostics.get('forecast_to_history_ratio'),
    }
    visuals = _merge_results_visual_defaults({
        'overview_cards': overview_cards,
        'diagnostics_rows': [{'metric': k.replace('_', ' ').title(), 'value': v} for k, v in diagnostics.items() if v is not None][:12],
        'model_ranking': model_ranking,
        'forecast_table': forecast_table,
        'charts': ([{'id': 'fallback-forecast-chart', 'title': 'Forecast Preview', 'subtitle': 'Fallback chart built from saved forecast output.', 'spec': json.dumps(chart_spec)}] if chart_spec else []),
        'render_warning': 'Detailed results visuals could not be rebuilt from the merged modeling table, so this page is showing a safe fallback view from the saved forecast output.'
    })
    return visuals

@router.get('/projects/{project_id}/results', response_class=HTMLResponse)
def results_page(project_id: int, request: Request, run_id: int, db: Session = Depends(get_db)):
    result_path = UPLOADS_DIR / f'project_{project_id}' / f'run_{run_id}_result.json'
    if not result_path.exists():
        return redirect(f'/projects/{project_id}/forecast?run_id={run_id}')
    result = json.loads(result_path.read_text(encoding='utf-8'))
    forecast_settings = _load_forecast_settings(project_id, run_id)
    refresh_visuals = str(request.query_params.get('refresh_visuals', '')).lower() in {'1', 'true', 'yes'}

    visuals = _normalize_cohort_maturity_labels(_normalize_company_total_payback_labels(_load_visuals_snapshot(project_id, run_id))) if not refresh_visuals else {}
    render_error = None
    snapshot_kind = str((visuals or {}).get('snapshot_kind') or '').lower()
    cached_has_renderable_charts = _snapshot_has_renderable_charts(visuals)
    cached_has_planning_payload = _snapshot_has_planning_payload(visuals)
    snapshot_is_minimal = snapshot_kind == 'minimal'
    mapping = _load_mapping_meta(project_id, run_id)
    snapshot_path = _project_dir(project_id) / f'run_{run_id}_visuals.json'
    dependency_paths = [result_path, _upload_meta_path(project_id, run_id), _mapping_meta_path(project_id, run_id)]
    cached_snapshot_stale = False
    try:
        if visuals and snapshot_path.exists():
            snapshot_mtime = snapshot_path.stat().st_mtime
            cached_snapshot_stale = any(path.exists() and path.stat().st_mtime > snapshot_mtime for path in dependency_paths)
    except Exception:
        cached_snapshot_stale = False
    cached_missing_mapped_cohort = bool(
        visuals
        and ((mapping or {}).get('cohort_revenue') or {}).get('file_key')
        and not visuals.get('cohort_rev_available')
    )
    roas_rows = (visuals or {}).get('cohort_rev_roas_table') or []
    cached_zero_cost_cohorts = bool(
        visuals
        and (visuals.get('cohort_rev_available') or roas_rows)
        and ((mapping or {}).get('cost') or {}).get('value_column')
        and roas_rows
        and all((row.get('total_cost') in (None, 0, 0.0, '0', '0.0')) for row in roas_rows[: min(len(roas_rows), 12)])
    )
    cached_placeholder_time_to_100 = bool(
        visuals
        and roas_rows
        and any(
            (
                row.get('time_to_100_roas_months') in (1, 1.0, '1')
                and row.get('roas') not in (None, '', 'None')
                and float(row.get('roas') or 0) < 1.0
            )
            for row in roas_rows[: min(len(roas_rows), 16)]
        )
    )
    cached_maturity_mismatch = bool(
        visuals
        and roas_rows
        and any(
            (
                ((lambda am: 'Mature' if am >= 6 else ('Growing' if am >= 3 else 'Early'))(
                    float(row.get('active_months'))
                )) != str(row.get('maturity_status') or '').strip()
            )
            for row in roas_rows[: min(len(roas_rows), 24)]
            if row.get('active_months') not in (None, '', 'None')
        )
    )
    cached_inflated_leads_chart = False
    if visuals and ((mapping or {}).get('leads') or {}).get('value_column'):
        try:
            for chart in (visuals.get('charts') or []):
                if chart.get('id') != 'results-rev-leads':
                    continue
                spec = json.loads(chart.get('spec') or '{}')
                for trace in (spec.get('data') or []):
                    if trace.get('name') != 'Leads':
                        continue
                    y_vals = [float(v) for v in (trace.get('y') or []) if v not in (None, '')]
                    if y_vals and max(y_vals) >= 100000:
                        cached_inflated_leads_chart = True
                    break
                break
        except Exception:
            cached_inflated_leads_chart = False
    cached_old_cohort_heatmap_palette = False
    if visuals and visuals.get('cohort_rev_heatmap_chart'):
        try:
            heatmap_spec = visuals.get('cohort_rev_heatmap_chart') or {}
            traces = heatmap_spec.get('data') or []
            if traces:
                colorscale = (traces[0] or {}).get('colorscale') or []
                color_values = [str(stop[1]).lower() for stop in colorscale if isinstance(stop, (list, tuple)) and len(stop) >= 2]
                cached_old_cohort_heatmap_palette = (
                    '#d58d5f' in color_values
                    or '#b22122' in color_values
                    or '#38bdf8' in color_values
                    or '#0f766e' in color_values
                )
        except Exception:
            cached_old_cohort_heatmap_palette = False
    should_rebuild_visuals = (
        refresh_visuals
        or not visuals
        or snapshot_is_minimal
        or not cached_has_renderable_charts
        or not cached_has_planning_payload
        or cached_snapshot_stale
        or cached_missing_mapped_cohort
        or cached_zero_cost_cohorts
        or cached_placeholder_time_to_100
        or cached_maturity_mismatch
        or cached_inflated_leads_chart
        or cached_old_cohort_heatmap_palette
    )
    if visuals and not should_rebuild_visuals:
        visuals = _normalize_cohort_maturity_labels(_normalize_company_total_payback_labels(_protect_results_visuals(visuals)))
        visuals['render_warning'] = visuals.get('render_warning') or 'Loaded saved run snapshot for a faster open. Refresh visuals to rebuild detailed sections from source files.'
    else:
        upload_meta = _load_upload_meta(project_id, run_id)
        try:
            df, build_error = _friendly_build_model_dataframe(upload_meta, mapping) if mapping is not None else (None, None)
            if mapping is not None and build_error:
                raise RuntimeError(build_error)
            visuals = _build_results_visuals(df, result, mapping=mapping, upload_meta=upload_meta)
            visuals['snapshot_kind'] = 'full'
            if not _snapshot_has_renderable_charts(visuals):
                raise RuntimeError('Full visuals build completed without any renderable chart specs.')
        except Exception as exc:
            render_error = str(exc)
            visuals = _build_minimal_results_visuals(result)
            visuals['snapshot_kind'] = 'minimal'
            visuals = _normalize_cohort_maturity_labels(_normalize_company_total_payback_labels(_protect_results_visuals(visuals)))
        if render_error:
            visuals['render_error'] = render_error
            if not visuals.get('render_warning'):
                visuals['render_warning'] = 'Detailed results visuals were partially unavailable, so safe defaults were applied for any missing sections.'
        try:
            _write_visuals_snapshot(project_id, run_id, visuals)
        except Exception:
            pass
    return templates.TemplateResponse('results.html', {
        'request': request,
        'project_id': project_id,
        'project': get_project(db, project_id),
        'run_id': run_id,
        'result': result,
        'visuals': visuals,
        'forecast_settings': forecast_settings,
    })


@router.get('/runs', response_class=HTMLResponse)
def runs_page(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse('saved_runs.html', {
        'request': request,
        'runs': list_runs(db),
        'projects': list_projects(db),
    })




@router.get('/runs/compare', response_class=HTMLResponse)
def compare_runs_page(
    request: Request,
    left_run_id: int | None = None,
    right_run_id: int | None = None,
    db: Session = Depends(get_db),
):
    runs = list_runs(db)
    compare_error = None
    left_run = get_run(db, left_run_id) if left_run_id else None
    right_run = get_run(db, right_run_id) if right_run_id else None

    left_project = get_project(db, left_run.project_id) if left_run else None
    right_project = get_project(db, right_run.project_id) if right_run else None

    left_result = _load_result_payload(left_run.project_id, left_run.id) if left_run else {}
    right_result = _load_result_payload(right_run.project_id, right_run.id) if right_run else {}

    left_summary = {}
    right_summary = {}
    visuals = {'charts': [], 'comparison_table': [], 'settings_rows': []}

    if left_run_id and right_run_id:
        if not left_run or not right_run:
            compare_error = 'One of the selected runs could not be found.'
        elif not left_result or not right_result:
            compare_error = 'Both runs need completed results before they can be compared side by side.'
        else:
            left_settings = _safe_run_settings(left_run.project_id, left_run.id)
            right_settings = _safe_run_settings(right_run.project_id, right_run.id)
            left_summary = _build_run_summary(left_project, left_run, left_result, left_settings)
            right_summary = _build_run_summary(right_project, right_run, right_result, right_settings)
            left_hist = _build_run_history_df(left_run.project_id, left_run.id)
            right_hist = _build_run_history_df(right_run.project_id, right_run.id)
            visuals = _build_compare_visuals(left_summary, right_summary, left_hist, right_hist, left_result, right_result)

    return templates.TemplateResponse('compare_runs.html', {
        'request': request,
        'runs': runs,
        'left_run_id': left_run_id,
        'right_run_id': right_run_id,
        'left_run': left_run,
        'right_run': right_run,
        'left_project': left_project,
        'right_project': right_project,
        'left_summary': left_summary,
        'right_summary': right_summary,
        'visuals': visuals,
        'compare_error': compare_error,
    })


@router.get('/projects/{project_id}/export/csv')
def export_csv(project_id: int, run_id: int):
    result_path = UPLOADS_DIR / f'project_{project_id}' / f'run_{run_id}_result.json'
    if not result_path.exists():
        raise HTTPException(status_code=404, detail='No results found for this run')
    payload = json.loads(result_path.read_text(encoding='utf-8'))
    df = pd.DataFrame(payload.get('ensemble_preview', []))
    stream = df.to_csv(index=False).encode('utf-8')
    return StreamingResponse(iter([stream]), media_type='text/csv', headers={'Content-Disposition': f'attachment; filename=run_{run_id}_forecast.csv'})


@router.get('/projects/{project_id}/export/xlsx')
def export_xlsx(project_id: int, run_id: int, db: Session = Depends(get_db)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    result_path = UPLOADS_DIR / f'project_{project_id}' / f'run_{run_id}_result.json'
    if not result_path.exists():
        raise HTTPException(status_code=404, detail='No results found for this run')

    payload = json.loads(result_path.read_text(encoding='utf-8'))
    project = get_project(db, project_id)
    run = get_run(db, run_id)
    settings = _load_forecast_settings(project_id, run_id)

    wb = Workbook()

    # ── helpers ──────────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill('solid', start_color='1F3864')
    ACCENT_FILL = PatternFill('solid', start_color='2E4A7A')
    ALT_FILL    = PatternFill('solid', start_color='F2F5FA')
    WHITE_FILL  = PatternFill('solid', start_color='FFFFFF')
    HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    BODY_FONT   = Font(name='Arial', size=10)
    BOLD_FONT   = Font(name='Arial', bold=True, size=10)
    TITLE_FONT  = Font(name='Arial', bold=True, size=14, color='1F3864')
    THIN        = Side(style='thin', color='C0C0C0')
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NUM_FMT     = '$#,##0.00'
    INT_FMT     = '#,##0'
    PCT_FMT     = '0.00%'

    def _style_header_row(ws, row_num: int, num_cols: int):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER

    def _style_data_row(ws, row_num: int, num_cols: int, alternate: bool = False):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = ALT_FILL if alternate else WHITE_FILL
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')

    def _auto_width(ws, min_w: int = 10, max_w: int = 30):
        for col_cells in ws.columns:
            length = max((len(str(c.value or '')) for c in col_cells), default=min_w)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)

    # ── Sheet 1: Forecast ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Forecast'
    ws1.sheet_view.showGridLines = False

    proj_name = project.name if project else f'Project {project_id}'
    run_label = run.label if run else f'Run {run_id}'
    ws1['A1'] = f'{proj_name} — Forecast Results'
    ws1['A1'].font = TITLE_FONT
    ws1['A2'] = f'Run: {run_label}'
    ws1['A2'].font = Font(name='Arial', italic=True, size=10, color='555555')
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[3].height = 18

    forecast_rows = payload.get('blended_forecast') or payload.get('ensemble_preview') or []
    forecast_df = pd.DataFrame(forecast_rows)

    if not forecast_df.empty:
        if 'expected' not in forecast_df.columns and 'yhat' in forecast_df.columns:
            forecast_df['expected'] = forecast_df['yhat']
        display_cols = [c for c in ['ds', 'expected', 'conservative', 'aggressive', 'lower', 'upper'] if c in forecast_df.columns]
        headers = {'ds': 'Date', 'expected': 'Expected ($)', 'conservative': 'Conservative ($)', 'aggressive': 'Aggressive ($)', 'lower': 'Lower Bound ($)', 'upper': 'Upper Bound ($)'}
        header_row = 4
        for ci, col in enumerate(display_cols, 1):
            ws1.cell(row=header_row, column=ci, value=headers.get(col, col))
        _style_header_row(ws1, header_row, len(display_cols))
        for ri, (_, row) in enumerate(forecast_df[display_cols].iterrows(), header_row + 1):
            alternate = (ri - header_row) % 2 == 0
            for ci, col in enumerate(display_cols, 1):
                val = row[col]
                cell = ws1.cell(row=ri, column=ci, value=str(val) if col == 'ds' else (_safe_number(val)))
                if col != 'ds' and cell.value is not None:
                    cell.number_format = NUM_FMT
            _style_data_row(ws1, ri, len(display_cols), alternate)
        # totals row
        total_row = header_row + len(forecast_df) + 1
        ws1.cell(row=total_row, column=1, value='TOTAL / AVG').font = BOLD_FONT
        ws1.cell(row=total_row, column=1).fill = ACCENT_FILL
        ws1.cell(row=total_row, column=1).font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        for ci, col in enumerate(display_cols[1:], 2):
            data_start = header_row + 1
            data_end = header_row + len(forecast_df)
            col_letter = get_column_letter(ci)
            ws1.cell(row=total_row, column=ci, value=f'=SUM({col_letter}{data_start}:{col_letter}{data_end})')
            ws1.cell(row=total_row, column=ci).number_format = NUM_FMT
            ws1.cell(row=total_row, column=ci).fill = ACCENT_FILL
            ws1.cell(row=total_row, column=ci).font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            ws1.cell(row=total_row, column=ci).border = BORDER
    else:
        ws1['A4'] = 'No forecast data available.'

    _auto_width(ws1)

    # ── Sheet 2: Model Rankings ───────────────────────────────────────────────
    ws2 = wb.create_sheet('Model Rankings')
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = 'Model Competition Results'
    ws2['A1'].font = TITLE_FONT
    ws2.row_dimensions[1].height = 22

    top_models = payload.get('top_models') or []
    rank_headers = ['Rank', 'Model', 'Final Score', 'MAE', 'RMSE', 'sMAPE', 'Bias']
    for ci, h in enumerate(rank_headers, 1):
        ws2.cell(row=3, column=ci, value=h)
    _style_header_row(ws2, 3, len(rank_headers))
    for ri, m in enumerate(top_models, 4):
        vals = [
            ri - 3,
            m.get('model') or m.get('model_name') or 'model',
            _safe_number(m.get('final_rank_score') or m.get('final_score')),
            _safe_number(m.get('mae')),
            _safe_number(m.get('rmse')),
            _safe_number(m.get('smape')),
            _safe_number(m.get('bias')),
        ]
        alternate = (ri - 3) % 2 == 0
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            if ci >= 3 and v is not None:
                cell.number_format = '0.0000'
        _style_data_row(ws2, ri, len(rank_headers), alternate)
    _auto_width(ws2)

    # ── Sheet 3: Diagnostics & Settings ──────────────────────────────────────
    ws3 = wb.create_sheet('Diagnostics')
    ws3.sheet_view.showGridLines = False
    ws3['A1'] = 'Run Diagnostics & Settings'
    ws3['A1'].font = TITLE_FONT
    ws3.row_dimensions[1].height = 22

    diag = payload.get('diagnostics') or {}
    hw = payload.get('hardware') or {}
    sett = payload.get('settings') or {}
    run_settings = settings or {}

    rows_data = [
        ('--- Diagnostics ---', ''),
        ('History Rows', diag.get('history_rows')),
        ('Recent Avg Revenue', diag.get('recent_avg_y')),
        ('Recent Median Revenue', diag.get('recent_median_y')),
        ('Recent Max Revenue', diag.get('recent_max_y')),
        ('Overall Avg Revenue', diag.get('overall_avg_y')),
        ('Forecast Avg Expected', diag.get('forecast_avg_expected')),
        ('Forecast / History Ratio', diag.get('forecast_to_history_ratio')),
        ('Avg Leads', diag.get('avg_leads')),
        ('Max Leads', diag.get('max_leads')),
        ('Avg Cost', diag.get('avg_cost')),
        ('Max Cost', diag.get('max_cost')),
        ('Duplicate DS Rows', diag.get('duplicate_ds_rows')),
        ('--- Hardware ---', ''),
        ('GPU Requested', str(hw.get('gpu_requested', False))),
        ('LightGBM Device', hw.get('lightgbm_device') or 'n/a'),
        ('Hardware Note', hw.get('note') or ''),
        ('--- Run Settings ---', ''),
        ('Business Profile', run_settings.get('business_profile') or ''),
        ('Frequency', run_settings.get('frequency') or ''),
        ('Horizon', run_settings.get('horizon') or ''),
        ('Scoring Metric', run_settings.get('scoring_metric') or ''),
        ('Output Mode', run_settings.get('output_mode') or ''),
        ('Revenue Lag Modeling', str(sett.get('revenue_lag_modeling_enabled', ''))),
        ('Revenue Lag Profile', sett.get('revenue_lag_profile') or ''),
        ('Feature Engineering', sett.get('feature_engineering') or ''),
    ]

    ws3.cell(row=3, column=1, value='Metric').font = HEADER_FONT
    ws3.cell(row=3, column=1).fill = HEADER_FILL
    ws3.cell(row=3, column=1).border = BORDER
    ws3.cell(row=3, column=2, value='Value').font = HEADER_FONT
    ws3.cell(row=3, column=2).fill = HEADER_FILL
    ws3.cell(row=3, column=2).border = BORDER
    ws3.column_dimensions['A'].width = 32
    ws3.column_dimensions['B'].width = 28

    ri = 4
    for label, value in rows_data:
        if str(label).startswith('---'):
            ws3.cell(row=ri, column=1, value=label.replace('-', '').strip()).font = Font(name='Arial', bold=True, size=10, color='1F3864')
            ws3.cell(row=ri, column=1).fill = ALT_FILL
            ws3.cell(row=ri, column=2).fill = ALT_FILL
        else:
            ws3.cell(row=ri, column=1, value=label).font = BODY_FONT
            cell_v = ws3.cell(row=ri, column=2, value=value)
            if isinstance(value, float) and value is not None:
                cell_v.number_format = NUM_FMT
            cell_v.font = BODY_FONT
        ws3.cell(row=ri, column=1).border = BORDER
        ws3.cell(row=ri, column=2).border = BORDER
        ri += 1

    # annotation note
    annotation_path = _project_dir(project_id) / f'run_{run_id}_annotation.json'
    if annotation_path.exists():
        ann = json.loads(annotation_path.read_text(encoding='utf-8'))
        ws3.cell(row=ri + 1, column=1, value='Run Notes').font = Font(name='Arial', bold=True, size=10, color='1F3864')
        ws3.cell(row=ri + 2, column=1, value=ann.get('note', '')).font = BODY_FONT
        ws3.cell(row=ri + 2, column=1).alignment = Alignment(wrap_text=True)
        ws3.merge_cells(start_row=ri + 2, start_column=1, end_row=ri + 2, end_column=2)
        ws3.row_dimensions[ri + 2].height = 60

    visuals = _load_visuals_snapshot(project_id, run_id)
    if not visuals:
        upload_meta = _load_upload_meta(project_id, run_id)
        mapping = _load_mapping_meta(project_id, run_id)
        try:
            df, build_error = _friendly_build_model_dataframe(upload_meta, mapping) if mapping is not None else (None, None)
            if mapping is not None and build_error:
                raise RuntimeError(build_error)
            visuals = _build_results_visuals(df, payload, mapping=mapping, upload_meta=upload_meta)
        except Exception:
            visuals = _build_minimal_results_visuals(payload)
            visuals = _normalize_cohort_maturity_labels(_normalize_company_total_payback_labels(_protect_results_visuals(visuals)))

    def _write_df_sheet(title: str, rows: list[dict], preferred_order: list[str] | None = None):
        if not rows:
            return
        ws = wb.create_sheet(title[:31])
        ws.sheet_view.showGridLines = False
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        df_sheet = pd.DataFrame(rows)
        if preferred_order:
            cols = [c for c in preferred_order if c in df_sheet.columns] + [c for c in df_sheet.columns if c not in preferred_order]
            df_sheet = df_sheet[cols]
        headers = list(df_sheet.columns)
        for ci, h in enumerate(headers, 1):
            ws.cell(row=3, column=ci, value=h)
        _style_header_row(ws, 3, len(headers))
        for ri2, (_, row2) in enumerate(df_sheet.iterrows(), 4):
            for ci, h in enumerate(headers, 1):
                val = row2[h]
                cell = ws.cell(row=ri2, column=ci, value=None if pd.isna(val) else val)
                if isinstance(val, (int, float, np.integer, np.floating)) and pd.notna(val):
                    cell.number_format = NUM_FMT if any(k in h.lower() for k in ['revenue', 'cost', 'spend']) else '0.00'
            _style_data_row(ws, ri2, len(headers), alternate=((ri2-3)%2==0))
        _auto_width(ws)

    _write_df_sheet('User Intelligence', visuals.get('whale_export_table') or [], ['whale_id','whale_tier','lifetime_revenue','recent_revenue','prior_revenue','delta_pct','revenue_share_pct','days_since_last_seen','status','action'])
    _write_df_sheet('Explainable AI', visuals.get('explainability_summary') or [], ['feature','importance','share_pct'])
    _write_df_sheet('LGBM SHAP Summary', visuals.get('lightgbm_shap_summary') or [], ['driver','feature','mean_abs_shap','share_pct'])
    _write_df_sheet('LGBM Rankings', visuals.get('lightgbm_ranked_opportunities') or [], ['dimension_type','dimension','rank','ranking_score','recommendation','business_reason','y','leads','cost','risk_score'])
    _write_df_sheet('Lead Quality', visuals.get('lightgbm_lead_quality_scores') or [], ['dimension_type','dimension','conversion_probability','high_value_probability','confidence_bucket','value_tier','lead_quality_summary'])
    _write_df_sheet('LGBM Anomalies', visuals.get('lightgbm_anomaly_rows') or [], ['date','series_id','actual_revenue','predicted_revenue','residual','anomaly_score','severity','likely_drivers','recommended_action'])
    _write_df_sheet('LGBM Whale Prediction', visuals.get('lightgbm_whale_predictions') or [], ['profile_id','behavior_label','cooling_risk','reactivation_likelihood','revenue_drop_risk','stability_score','lifetime_revenue','recent_revenue','recommended_action'])
    _write_df_sheet('LGBM Response Curve', visuals.get('lightgbm_response_curve') or [], ['spend_change_pct','projected_revenue','projected_roas','marginal_roas','zone','business_label'])
    _write_df_sheet('LGBM Simulations', visuals.get('lightgbm_scenarios') or [], ['scenario','projected_revenue','projected_roas','projected_lead_quality_effect','risk_band_low','risk_band_high','interpretation','warning'])
    _write_df_sheet('Affiliate Quality', visuals.get('lightgbm_affiliate_quality') or [], ['affiliate','quality_score','quality_tier','recommendation','recent_revenue','trend_pct','consistency_score','short_term_strong_long_term_weak'])
    _write_df_sheet('LGBM Confidence', visuals.get('lightgbm_confidence_rows') or [], ['surface','risk_label','confidence_low','confidence_high','reason'])
    _write_df_sheet('Attribution', visuals.get('attribution_table') or [], ['channel','revenue','revenue_share_pct','roas','recent_roas','rev_per_lead','total_cost','stability_score'])
    _write_df_sheet('Scenario Simulator', visuals.get('scenario2_table') or [], ['scenario','projected_spend','projected_leads','projected_revenue','projected_roas'])
    _write_df_sheet('Cohorts', visuals.get('cohort_rev_table') or [], None)

    report_ws = wb.create_sheet('Auto-Report')
    report_ws.sheet_view.showGridLines = False
    report_ws['A1'] = 'Auto-Report'
    report_ws['A1'].font = TITLE_FONT
    for idx, line in enumerate(visuals.get('stakeholder_report_lines') or ['No auto-report available for this run.'], start=3):
        report_ws.cell(row=idx, column=1, value=f'• {line}')
        report_ws.cell(row=idx, column=1).alignment = Alignment(wrap_text=True)
    report_ws.column_dimensions['A'].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'forecast_run_{run_id}_{proj_name.replace(" ", "_")}.xlsx'
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename}'})


@router.get('/projects/{project_id}/export/report_html')
def export_report_html(project_id: int, run_id: int):
    result_path = UPLOADS_DIR / f'project_{project_id}' / f'run_{run_id}_result.json'
    if not result_path.exists():
        raise HTTPException(status_code=404, detail='No results found for this run')
    payload = json.loads(result_path.read_text(encoding='utf-8'))
    visuals = _load_visuals_snapshot(project_id, run_id)
    if not visuals:
        upload_meta = _load_upload_meta(project_id, run_id)
        mapping = _load_mapping_meta(project_id, run_id)
        try:
            df, build_error = _friendly_build_model_dataframe(upload_meta, mapping) if mapping is not None else (None, None)
            if mapping is not None and build_error:
                raise RuntimeError(build_error)
            visuals = _build_results_visuals(df, payload, mapping=mapping, upload_meta=upload_meta)
        except Exception:
            visuals = _build_minimal_results_visuals(payload)
    visuals = _protect_results_visuals(visuals)
    lines = visuals.get('stakeholder_report_lines') or ['No auto-report available for this run.']
    html_lines = ''.join(f'<li>{html.escape(str(line))}</li>' for line in lines)
    doc = (
        '<!doctype html><html><head><meta charset="utf-8"><title>Auto-Report</title>'
        '<style>body{font-family:Segoe UI,Arial,sans-serif;background:#f5f7fb;color:#122033;padding:32px}'
        '.card{max-width:900px;margin:0 auto;background:#fff;border:1px solid #dbe3f0;border-radius:16px;padding:24px;box-shadow:0 10px 30px rgba(16,24,40,.08)}'
        'h1{margin-top:0}li{margin:0 0 12px 0;line-height:1.5}</style></head><body>'
        f'<div class="card"><h1>Auto-Report</h1><ul>{html_lines}</ul></div></body></html>'
    )
    return StreamingResponse(iter([doc.encode('utf-8')]), media_type='text/html', headers={'Content-Disposition': f'attachment; filename=forecast_run_{run_id}_auto_report.html'})


# ── Annotation routes ─────────────────────────────────────────────────────────

def _annotation_path(project_id: int, run_id: int) -> Path:
    return _project_dir(project_id) / f'run_{run_id}_annotation.json'

def _project_storage_dir(project_id: int) -> Path:
    path = PROJECTS_DIR / f'project_{project_id}'
    path.mkdir(parents=True, exist_ok=True)
    return path


def _scenario_dir(project_id: int) -> Path:
    path = _project_storage_dir(project_id) / 'scenarios'
    path.mkdir(parents=True, exist_ok=True)
    return path


def _scenario_slug(name: str) -> str:
    raw = str(name or '').strip().lower()
    slug = re.sub(r'[^a-z0-9]+', '_', raw).strip('_')
    return slug[:80] or 'scenario'


def _scenario_file_path(project_id: int, run_id: int, slug: str) -> Path:
    safe_slug = _scenario_slug(slug)
    return _scenario_dir(project_id) / f'run_{run_id}__{safe_slug}.json'


def _list_saved_scenarios(project_id: int, run_id: int) -> list[dict]:
    out = []
    for path in sorted(_scenario_dir(project_id).glob(f'run_{run_id}__*.json')):
        try:
            item = json.loads(path.read_text(encoding='utf-8'))
            if not isinstance(item, dict):
                continue
            item['id'] = path.stem.replace(f'run_{run_id}__', '', 1)
            item['filename'] = path.name
            out.append(item)
        except Exception:
            continue
    out.sort(key=lambda x: str(x.get('updated_at') or x.get('savedAt') or ''), reverse=True)
    return out



@router.get('/api/projects/{project_id}/runs/{run_id}/scenarios')
def get_saved_scenarios(project_id: int, run_id: int):
    return JSONResponse({'items': _list_saved_scenarios(project_id, run_id)})


@router.post('/api/projects/{project_id}/runs/{run_id}/scenarios')
async def save_scenario(project_id: int, run_id: int, request: Request):
    body = await request.json()
    name = str(body.get('name', '')).strip()[:120] or 'Scenario'
    slug = _scenario_slug(body.get('id') or name)
    budget = _safe_number(body.get('budget')) or 0.0
    revenue = _safe_number(body.get('revenue')) or 0.0
    delta = _safe_number(body.get('delta')) or 0.0
    roas = _safe_number(body.get('roas')) or 0.0
    record = {
        'id': slug,
        'name': name,
        'project_id': project_id,
        'run_id': run_id,
        'budget': budget,
        'revenue': revenue,
        'delta': delta,
        'roas': roas,
        'savedAt': body.get('savedAt') or pd.Timestamp.utcnow().isoformat(),
        'updated_at': pd.Timestamp.utcnow().isoformat(),
    }
    path = _scenario_file_path(project_id, run_id, slug)
    path.write_text(json.dumps(record, indent=2), encoding='utf-8')
    return JSONResponse({'ok': True, 'item': record, 'items': _list_saved_scenarios(project_id, run_id)})


@router.delete('/api/projects/{project_id}/runs/{run_id}/scenarios/{scenario_id}')
def delete_saved_scenario(project_id: int, run_id: int, scenario_id: str):
    path = _scenario_file_path(project_id, run_id, scenario_id)
    removed = False
    try:
        if path.exists():
            path.unlink()
            removed = True
    except Exception:
        removed = False
    return JSONResponse({'ok': removed, 'items': _list_saved_scenarios(project_id, run_id)})


@router.get('/api/projects/{project_id}/runs/{run_id}/annotation')
def get_annotation(project_id: int, run_id: int):
    path = _annotation_path(project_id, run_id)
    if path.exists():
        return JSONResponse(json.loads(path.read_text(encoding='utf-8')))
    return JSONResponse({'note': '', 'tags': [], 'updated_at': None})


@router.post('/api/projects/{project_id}/runs/{run_id}/annotation')
async def save_annotation(project_id: int, run_id: int, request: Request):
    body = await request.json()
    note = str(body.get('note', ''))[:2000]
    tags_raw = body.get('tags', [])
    tags = [str(t).strip() for t in tags_raw if str(t).strip()][:10]
    data = {'note': note, 'tags': tags, 'updated_at': pd.Timestamp.utcnow().isoformat()}
    path = _annotation_path(project_id, run_id)
    path.write_text(json.dumps(data), encoding='utf-8')
    return JSONResponse({'ok': True, 'note': note, 'tags': tags})
